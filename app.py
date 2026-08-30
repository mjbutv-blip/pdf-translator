import io
import json
import os
import zipfile

import fitz
import openpyxl
import pandas as pd
import streamlit as st
from openai import OpenAI


def _prime_streamlit_secrets_env() -> None:
    for name in (
        "DATABASE_URL",
        "OPENAI_API_KEY",
        "ANTHROPIC_API_KEY",
        "OPENAI_MODEL",
        "OPENAI_TIMEOUT_SECONDS",
        "OPENAI_REASONING_EFFORT",
        "OPENAI_FALLBACK_MODELS",
        "PDF_TRANSLATION_BATCH_SIZE",
        "RFQ_MANAGER_API_URL",
    ):
        if os.environ.get(name):
            continue
        try:
            value = st.secrets.get(name, "")
        except Exception:
            value = ""
        if value:
            os.environ[name] = str(value)


_prime_streamlit_secrets_env()

from config import DEFAULT_FONT, DEFAULT_GLOSSARY, _use_postgres
from translation_core import *
from translation_core import _load_default_glossary_df
from translation_jobs import *
from translation_jobs import _RUNNING_JOB_THREADS

@st.cache_resource(show_spinner=False)
def run_startup_tasks_once(use_postgres_marker: bool, schema_version: int) -> bool:
    init_db()
    seed_demo_data_if_empty()
    sync_staff_customer_assignments()
    return True


def render_login_panel() -> dict | None:
    current_user = st.session_state.get("current_user")
    if current_user:
        with st.sidebar:
            st.subheader("当前用户")
            st.write(f"用户名：**{current_user['username']}**")
            st.write(f"角色：`{current_user['role']}`")
            if current_user.get("group_name"):
                st.write(f"小组：**{current_user['group_name']}**")

            accessible = get_accessible_customers(current_user)
            st.caption(f"可访问客户：{len(accessible)} 个")
            if accessible:
                st.dataframe(
                    pd.DataFrame(accessible)[
                        ["customer_code", "customer_name", "group_name", "assigned_staff_username"]
                    ],
                    use_container_width=True,
                    hide_index=True,
                )

            if st.button("退出登录", use_container_width=True):
                for key in ("current_user", "selected_customer_id"):
                    st.session_state.pop(key, None)
                st.rerun()
        return current_user

    st.subheader("登录")
    st.caption("第一版本地账号，用于验证客户术语库权限。")
    with st.form("login_form"):
        username = st.text_input("用户名")
        password = st.text_input("密码", type="password")
        submitted = st.form_submit_button("登录", use_container_width=True, type="primary")

    if submitted:
        user = authenticate_user(username, password)
        if user:
            st.session_state["current_user"] = user
            st.rerun()
        st.error("用户名或密码不正确")

    st.info(
        "测试账号：admin/admin123，leader_a/leader123，"
        "staff_1/staff123，staff_2/staff123，staff_3/staff123"
    )
    return None


def render_customer_selector(user: dict) -> str | None:
    accessible = get_accessible_customers(user)
    valid_ids = [c["customer_id"] for c in accessible]
    pending_selected = st.session_state.pop("_pending_selected_customer_id", None)
    if pending_selected in valid_ids:
        st.session_state["selected_customer_id"] = pending_selected
    selected = st.session_state.get("selected_customer_id")
    if selected not in valid_ids:
        st.session_state.pop("selected_customer_id", None)
        selected = None

    st.subheader("客户选择")
    if not accessible:
        st.warning("当前用户没有可访问客户。")
        return None

    labels = {
        c["customer_id"]: (
            f"{c['customer_code']} / {c['customer_name']} / "
            f"{c['group_name']} / {c['assigned_staff_username']}"
        )
        for c in accessible
    }
    default_index = valid_ids.index(selected) if selected in valid_ids else 0
    selected_customer_id = st.selectbox(
        "当前客户",
        valid_ids,
        index=default_index,
        format_func=lambda cid: labels.get(cid, cid),
        key="selected_customer_id",
    )
    st.caption("后续 PDF / Excel 翻译会使用这里选择的客户术语库。")
    return selected_customer_id


# ── Glossary helpers ───────────────────────────────────────────────────────────

def render_worker_health_indicator() -> None:
    health = get_worker_queue_health()
    live_workers = int(health.get("live_worker_count") or 0)
    queued = int(health.get("queued_count") or 0)
    running = int(health.get("running_count") or 0)
    queue_parts = (
        f"等待任务：{queued}（PDF {health.get('queued_pdf_count', 0)} / Excel {health.get('queued_excel_count', 0)}）"
        f"｜正在翻译：{running}（PDF {health.get('running_pdf_count', 0)} / Excel {health.get('running_excel_count', 0)}）"
    )
    if live_workers <= 0 and queued > 0:
        st.warning(
            "后台翻译服务暂未检测到。当前有等待中的翻译任务，这些任务暂时不会开始翻译，"
            "直到后台翻译服务启动。"
        )
        st.caption(queue_parts)
        return
    if live_workers <= 0:
        st.info("后台翻译服务暂未检测到。新任务可以提交，但会保持等待状态，直到后台翻译服务启动。")
        return
    if queued or running:
        st.success(f"后台翻译服务运行中（{live_workers} 个 worker）。{queue_parts}")
    else:
        st.success(f"后台翻译服务运行正常（{live_workers} 个 worker）。")


@st.fragment
def render_pdf_jobs_panel(current_user: dict, api_key: str) -> None:
    pdf_jobs = list_translation_jobs(current_user["username"], "PDF") if current_user else []
    if pdf_jobs and api_key and os.getenv("PDF_WORKER_MODE", "external").strip().lower() == "embedded":
        has_active_pdf_thread = any(
            thread.is_alive() and not is_translation_job_cancelled(job_id)
            for job_id, thread in _RUNNING_JOB_THREADS.items()
        )
        has_running_pdf_job = any(job["status"] == "running" for job in pdf_jobs)
        queued_pdf_job = next((job for job in reversed(pdf_jobs) if job["status"] == "queued"), None)
        if queued_pdf_job and not has_active_pdf_thread and not has_running_pdf_job:
            start_pdf_translation_job(
                queued_pdf_job["job_id"],
                api_key,
                start_next_on_finish=True,
            )
            st.rerun(scope="fragment")
    if os.getenv("PDF_WORKER_MODE", "external").strip().lower() != "embedded":
        render_worker_health_indicator()
    if not pdf_jobs:
        return

    st.divider()
    st.subheader("PDF 后台任务")
    if os.getenv("PDF_WORKER_MODE", "external").strip().lower() != "embedded":
        st.caption("PDF 后台翻译由独立 worker 处理。请确保 worker.py 正在运行。")
    refresh_col, cancel_col = st.columns(2)
    with refresh_col:
        if st.button("刷新任务状态", use_container_width=True, key="refresh_pdf_jobs_btn"):
            st.rerun(scope="fragment")
    with cancel_col:
        if st.button("取消等待/翻译中的 PDF", use_container_width=True, key="cancel_running_pdf_jobs_btn"):
            cancel_result = cancel_pdf_translation_jobs(current_user["username"])
            if cancel_result["cancelled_count"]:
                st.success(f"已取消 {cancel_result['cancelled_count']} 个 PDF 任务。")
            else:
                st.info("当前没有等待中或翻译中的 PDF 任务。")
            st.rerun(scope="fragment")
    for job in pdf_jobs:
        label = {
            "queued": "等待中",
            "running": "翻译中",
            "complete": "已完成",
            "failed": "失败",
        }.get(job["status"], job["status"])
        with st.expander(f"{label} · {job['source_file_name']} · {job['updated_at']}", expanded=job["status"] in {"running", "failed"}):
            st.progress(float(job.get("progress") or 0), text=job.get("message") or label)
            if job.get("error") and job.get("error") != PDF_JOB_CANCELLED_ERROR:
                st.error(job["error"])
            if job.get("error") == PDF_JOB_CANCELLED_ERROR:
                st.info("该任务已取消。可以重新上传文件开始翻译。")
            if job["status"] == "failed":
                delete_translation_job(job["job_id"], current_user["username"])
            if job["status"] == "running":
                if os.getenv("PDF_WORKER_MODE", "external").strip().lower() == "embedded":
                    st.caption("如果进度长时间不动，可以重新启动这个后台任务。")
                    if st.button(
                        "重新启动任务",
                        use_container_width=True,
                        key=f"restart_pdf_job_{job['job_id']}",
                    ):
                        restarted = start_pdf_translation_job(
                            job["job_id"],
                            api_key,
                            start_next_on_finish=True,
                        )
                        if restarted:
                            st.success("已重新启动任务。")
                        else:
                            st.info("任务线程仍在运行，请稍后刷新查看。")
                        st.rerun(scope="fragment")
                else:
                    st.caption("该任务由独立 PDF worker 执行；如长时间无变化，请检查 worker.py 是否运行。")
            if job["status"] == "complete":
                full_job = get_translation_job(job["job_id"], current_user["username"])
                if full_job:
                    meta = json.loads(full_job.get("result_meta") or "{}")
                    base = full_job["source_file_name"].rsplit(".", 1)[0]
                    review_summary = meta.get("review_summary") or {}
                    if review_summary.get("summary_text"):
                        st.info(review_summary["summary_text"])
                    else:
                        st.caption(f"未收录术语：{meta.get('n_terms', 0)} 条")
                    if review_summary.get("needs_manual_review") and review_summary.get("review_locations"):
                        st.caption(f"重点核查位置：{'、'.join(review_summary['review_locations'][:3])}")
                    dl_cols = st.columns(2)
                    with dl_cols[0]:
                        st.download_button(
                            "⬇️ 下载中文 PDF",
                            data=full_job["result_file"],
                            file_name=f"{base}_translated.pdf",
                            mime="application/pdf",
                            use_container_width=True,
                            key=f"job_pdf_dl_{job['job_id']}",
                        )
                    with dl_cols[1]:
                        if full_job.get("result_report"):
                            report_kind = meta.get("report_kind") or "xlsx"
                            report_ext = "zip" if report_kind == "zip" else "xlsx"
                            report_mime = "application/zip" if report_kind == "zip" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            st.download_button(
                                "⬇️ 下载报告",
                                data=full_job["result_report"],
                                file_name=f"{base}_report.{report_ext}",
                                mime=report_mime,
                                use_container_width=True,
                                key=f"job_pdf_report_dl_{job['job_id']}",
                            )


# ── Streamlit page ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="服装行业翻译引擎",
    page_icon="🧵",
    layout="centered",
    initial_sidebar_state="expanded",
)

st.title("🧵 服装行业翻译引擎")
st.caption("支持 PDF 与 Excel (.xlsx) 双格式 · 上传文件 + 术语库 · 调用 OpenAI 自动翻译为中文")
st.divider()

run_startup_tasks_once(_use_postgres(), 1)
current_user = render_login_panel()
if not current_user:
    st.stop()
selected_customer_id = render_customer_selector(current_user)
if selected_customer_id:
    selected_customer = get_customer(selected_customer_id)
    if selected_customer:
        st.info(
            f"当前选择客户：**{selected_customer['customer_code']} / "
            f"{selected_customer['customer_name']}**（{selected_customer['group_name']}）"
        )

# ── 会话内术语库初始化 ─────────────────────────────────────────────────────────
if "glossary_df" not in st.session_state:
    st.session_state["glossary_df"] = _load_default_glossary_df()
    st.session_state["glossary_source"] = (
        DEFAULT_GLOSSARY.name if DEFAULT_GLOSSARY.exists() else "（空，未找到默认术语库）"
    )
    st.session_state["glossary_conflicts"] = pd.DataFrame(columns=["英文术语", "旧翻译", "新翻译（采用）"])

# ── API Key（全局共用）───────────────────────────────────────────────────────
api_key = st.text_input(
    "🔑 OpenAI API Key",
    type="password",
    value=os.environ.get("OPENAI_API_KEY", "") or os.environ.get("ANTHROPIC_API_KEY", ""),
    placeholder="sk-...",
)
st.divider()

tab_labels = ["📄 PDF 翻译（支持批量）", "📊 Excel 翻译", "📚 客户术语库", "🧩 术语候选"]
if can_approve_glossary_change(current_user):
    tab_labels.append("✅ 术语审批")
    tab_labels.append("👤 用户管理")
    tab_labels.append("🏢 客户管理")
tabs = st.tabs(tab_labels)
tab_pdf, tab_excel, tab_glossary, tab_candidates = tabs[:4]
tab_approval = tabs[4] if len(tabs) > 4 else None
tab_user_admin = tabs[5] if len(tabs) > 5 else None
tab_customer_admin = tabs[6] if len(tabs) > 6 else None

# ════════════════════════════════════════════════════════════════════════════
#  PDF Tab — batch upload, independent processing per file
# ════════════════════════════════════════════════════════════════════════════
with tab_pdf:
    if "pdf_upload_reset_nonce" not in st.session_state:
        st.session_state["pdf_upload_reset_nonce"] = 0
    if st.button(
        "取消当前 PDF 翻译并重新上传",
        use_container_width=True,
        key="cancel_reset_pdf_jobs_btn",
    ):
        cancel_result = cancel_pdf_translation_jobs(current_user["username"])
        st.session_state.pop("last_pdf_job_ids", None)
        st.session_state.pop("pdf_batch_results", None)
        st.session_state["pdf_upload_reset_nonce"] += 1
        if cancel_result["cancelled_count"]:
            st.success(f"已取消 {cancel_result['cancelled_count']} 个 PDF 任务，可以重新上传。")
        else:
            st.info("当前没有等待中或翻译中的 PDF 任务，已重置上传区域。")
        st.rerun()

    pdf_upload_nonce = st.session_state["pdf_upload_reset_nonce"]
    pdf_files = st.file_uploader(
        "上传待翻译 PDF（可一次选择多个文件）",
        type=["pdf"],
        accept_multiple_files=True,
        key=f"pdf_uploader_{pdf_upload_nonce}",
    )
    font_label = (
        f"🔤 上传中文字体 TTF（可选，已检测到默认字体 {DEFAULT_FONT.name}）"
        if DEFAULT_FONT.exists()
        else "🔤 上传中文字体 TTF（必填，未检测到默认字体）"
    )
    font_file = st.file_uploader(font_label, type=["ttf"], key=f"pdf_font_uploader_{pdf_upload_nonce}")
    font_ready = bool(font_file or DEFAULT_FONT.exists())
    selected_glossary_count = len(get_customer_glossary_df(selected_customer_id)) if selected_customer_id else 0
    st.caption(f"📚 当前客户术语库：**{selected_customer_id or '未选择'}** · {selected_glossary_count} 条 active 术语")

    pdf_scope_choice = st.radio(
        "翻译范围",
        ["all", "workmanship_auto", "manual"],
        format_func=lambda v: {
            "all": "全部页面",
            "workmanship_auto": "自动识别做工页",
            "manual": "手动选择页面",
        }[v],
        horizontal=True,
        key="pdf_scope_choice",
    )
    pdf_scope_configs: dict[str, dict] = {}
    pdf_scope_errors: list[str] = []
    if pdf_files and pdf_scope_choice != "all":
        st.caption("页码请按 PDF 阅读器里的页码填写，从 1 开始，例如：1-7,10。")
        for file_idx, pf in enumerate(pdf_files):
            pdf_bytes_for_scope = pf.getvalue()
            detection_rows = []
            try:
                if pdf_scope_choice == "workmanship_auto":
                    detection_rows = detect_workmanship_pages(pdf_bytes_for_scope)
                    total_pages = len(detection_rows)
                    default_pages = [
                        row["page_number"] for row in detection_rows
                        if row["is_workmanship"]
                    ]
                else:
                    doc_for_count = fitz.open(stream=pdf_bytes_for_scope, filetype="pdf")
                    total_pages = len(doc_for_count)
                    doc_for_count.close()
                    default_pages = list(range(1, total_pages + 1))
                default_range = format_page_ranges(default_pages)
                with st.expander(f"翻译范围确认：{pf.name}", expanded=True):
                    if detection_rows:
                        preview_df = pd.DataFrame([{
                            "页码": row["page_number"],
                            "判断": "做工" if row["is_workmanship"] else "跳过",
                            "分数": row["score"],
                            "原因": row["reason"],
                        } for row in detection_rows])
                        st.dataframe(preview_df, use_container_width=True, hide_index=True, height=240)
                    page_range = st.text_input(
                        "本文件实际翻译页码",
                        value=default_range,
                        key=f"pdf_scope_pages_{file_idx}_{pf.name}",
                    )
                selected_pages, page_errors = parse_page_ranges(page_range, total_pages)
                if page_errors:
                    pdf_scope_errors.extend([f"{pf.name}：{err}" for err in page_errors])
                if not selected_pages:
                    pdf_scope_errors.append(f"{pf.name}：至少选择 1 页。")
                pdf_scope_configs[pf.name] = {
                    "selected_pages": selected_pages,
                    "detection": detection_rows,
                    "total_pages": total_pages,
                    "page_range": page_range,
                }
            except Exception as exc:
                pdf_scope_errors.append(f"{pf.name}：范围识别失败：{exc}")
        for err in pdf_scope_errors:
            st.error(err)

    customer_ready = bool(selected_customer_id and can_use_customer_glossary(current_user, selected_customer_id))
    can_start_pdf = bool(pdf_files and api_key and font_ready and customer_ready and not pdf_scope_errors)
    if not can_start_pdf:
        missing = []
        if not pdf_files:
            missing.append("PDF 文件")
        if not api_key:
            missing.append("OpenAI API Key")
        if not font_ready:
            missing.append("中文字体 TTF")
        if not customer_ready:
            missing.append("有权限的客户")
        st.info(f"请先提供：{'、'.join(missing)}")

    start_pdf_btn = st.button(
        "🚀  开始翻译 PDF",
        disabled=not can_start_pdf,
        use_container_width=True,
        type="primary",
        key="start_pdf_btn",
    )

    if start_pdf_btn:
        created_jobs = []
        font_bytes = font_file.getvalue() if font_file else None
        pdf_execution_mode = (
            "embedded"
            if os.getenv("PDF_WORKER_MODE", "external").strip().lower() == "embedded"
            else "external"
        )
        for pf in pdf_files:
            pdf_input_bytes = pf.getvalue()
            scope_cfg = pdf_scope_configs.get(pf.name, {})
            if pdf_scope_choice == "all":
                doc_for_count = fitz.open(stream=pdf_input_bytes, filetype="pdf")
                scope_cfg = {"total_pages": len(doc_for_count), "detection": []}
                doc_for_count.close()
            selected_pages = None if pdf_scope_choice == "all" else scope_cfg.get("selected_pages")
            scope_detection = scope_cfg.get("detection") if pdf_scope_choice != "all" else []
            job_id = create_translation_job(
                job_type="PDF",
                username=current_user["username"],
                customer_id=selected_customer_id,
                source_file_name=pf.name,
                input_bytes=pdf_input_bytes,
                aux_bytes=font_bytes,
                config={
                    "scope_mode": pdf_scope_choice,
                    "selected_pages": selected_pages,
                    "scope_detection": scope_detection,
                    "scope_cfg": scope_cfg,
                },
                execution_mode=pdf_execution_mode,
            )
            created_jobs.append(job_id)
        if created_jobs and os.getenv("PDF_WORKER_MODE", "external").strip().lower() == "embedded":
            start_pdf_translation_job(created_jobs[0], api_key, start_next_on_finish=True)
        st.session_state["last_pdf_job_ids"] = created_jobs
        st.success(f"已创建 **{len(created_jobs)}** 个后台翻译任务。可以切换页面，稍后回到本页下载结果。")
        if os.getenv("PDF_WORKER_MODE", "external").strip().lower() != "embedded":
            st.info("任务已进入等待队列，将由独立 PDF worker 执行。")

    render_pdf_jobs_panel(current_user, api_key)

    batch_results = st.session_state.get("pdf_batch_results", [])
    if batch_results:
        st.divider()
        n_ok = sum(1 for r in batch_results if r["ok"])
        st.success(f"批量处理完成：成功 **{n_ok}** / 失败 **{len(batch_results) - n_ok}**（共 {len(batch_results)} 个文件）")

        # 全部成功结果打包为 ZIP
        ok_results = [r for r in batch_results if r["ok"]]
        if len(ok_results) > 1:
            zip_buf = io.BytesIO()
            with zipfile.ZipFile(zip_buf, "w") as zf:
                for r in ok_results:
                    base = r["name"].rsplit(".", 1)[0]
                    zf.writestr(f"{base}_translated.pdf", r["pdf"])
                    if r["n_terms"] > 0:
                        zf.writestr(f"{base}_unrecorded_terms.xlsx", r["xlsx"])
                    if r.get("scope_report"):
                        zf.writestr(
                            f"{base}_翻译范围报告.xlsx",
                            build_scope_report_xlsx(r["scope_report"]),
                        )
            st.download_button(
                label=f"⬇️  打包下载全部 {len(ok_results)} 个译文（ZIP）",
                data=zip_buf.getvalue(),
                file_name="translated_pdfs.zip",
                mime="application/zip",
                use_container_width=True,
                type="primary",
            )

        for i, r in enumerate(batch_results):
            with st.expander(f"{'✅' if r['ok'] else '❌'} {r['name']}", expanded=not r["ok"]):
                if r["ok"]:
                    base = r["name"].rsplit(".", 1)[0]
                    st.caption(f"未收录术语：{r['n_terms']} 条")
                    dl1, dl2, dl3 = st.columns(3)
                    with dl1:
                        st.download_button(
                            label="⬇️  下载中文 PDF",
                            data=r["pdf"],
                            file_name=f"{base}_translated.pdf",
                            mime="application/pdf",
                            use_container_width=True,
                            key=f"pdf_dl_{i}",
                        )
                    with dl2:
                        if r["n_terms"] > 0:
                            st.download_button(
                                label="⬇️  下载未收录术语",
                                data=r["xlsx"],
                                file_name=f"{base}_unrecorded_terms.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key=f"xlsx_dl_{i}",
                            )
                    with dl3:
                        if r.get("scope_report"):
                            st.download_button(
                                label="⬇️ 下载范围报告",
                                data=build_scope_report_xlsx(r["scope_report"]),
                                file_name=f"{base}_翻译范围报告.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key=f"pdf_scope_report_dl_{i}",
                            )
                else:
                    st.error(r["error"])

# ════════════════════════════════════════════════════════════════════════════
#  Excel Tab
# ════════════════════════════════════════════════════════════════════════════
def _build_excel_report_bytes(report_rows: list[dict], img_report_rows: list[dict] | None = None) -> bytes:
    report_wb = openpyxl.Workbook()
    report_ws = report_wb.active
    report_ws.title = "翻译报告"
    headers = ["sheet_name", "cell_coordinate", "original_text", "translated_text",
               "status", "skip_reason", "is_merged_cell", "layout_warning",
               "scope_mode", "selected_sheets", "skipped_sheets",
               "detection_score", "detection_reason"]
    report_ws.append(headers)
    for cell in report_ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for r in report_rows:
        report_ws.append([r.get(h, "") for h in headers])
    for col_letter, width in zip("ABCDEFGHIJKLM", [12, 14, 35, 35, 12, 24, 14, 14, 18, 35, 35, 12, 70]):
        report_ws.column_dimensions[col_letter].width = width

    if img_report_rows:
        img_ws = report_wb.create_sheet("图片译文")
        img_headers = ["drawing", "image", "status", "original_text", "translated_text", "skip_reason"]
        img_ws.append(img_headers)
        for cell in img_ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        for r in img_report_rows:
            img_ws.append([r.get(h, "") for h in img_headers])
        for col_letter, width in zip("ABCDEF", [25, 18, 10, 35, 35, 20]):
            img_ws.column_dimensions[col_letter].width = width

    buf = io.BytesIO()
    report_wb.save(buf)
    return buf.getvalue()


def render_customer_glossary_import_panel(
    user: dict,
    customer_id: str | None,
    key_prefix: str,
) -> None:
    st.subheader("上传/更新客户术语库")
    if not customer_id:
        st.info("请先选择客户。")
        return
    if not can_submit_glossary_change(user, customer_id):
        st.warning("当前用户无权为该客户上传术语库。")
        return

    mode_labels = {
        "skip_existing": "跳过已存在术语",
        "overwrite_existing": "覆盖已有术语",
        "pending_request": "生成待审批申请",
    }
    if can_approve_glossary_change(user):
        conflict_mode = st.radio(
            "遇到同一客户下已存在的英文术语时",
            ["skip_existing", "overwrite_existing", "pending_request"],
            format_func=lambda v: mode_labels[v],
            horizontal=True,
            key=f"{key_prefix}_conflict_mode",
        )
    else:
        conflict_mode = "pending_request"
        st.caption("当前角色上传后会生成 pending 术语申请，等待公司管理员审批。")

    uploaded = st.file_uploader(
        "上传术语库 Excel",
        type=["xlsx", "xls"],
        key=f"{key_prefix}_glossary_upload",
    )
    if not uploaded:
        return

    df, missing, parse_stats, parse_report = parse_customer_glossary_excel(uploaded.getvalue(), uploaded.name)
    if missing:
        st.error("；".join(missing))
        if parse_report:
            st.download_button(
                "⬇️ 下载导入错误报告",
                data=build_import_report_xlsx(parse_report),
                file_name="import_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"{key_prefix}_parse_report_dl",
            )
        return

    st.caption(
        f"读取行数 **{parse_stats.get('total_rows', 0)}** ｜ "
        f"有效行 **{parse_stats.get('valid_rows', 0)}** ｜ "
        f"跳过空行 **{parse_stats.get('skipped_blank', 0)}** ｜ "
        f"文件内重复 **{parse_stats.get('duplicate_terms', 0)}** ｜ "
        f"错误行 **{parse_stats.get('error_rows', 0)}**"
    )
    st.dataframe(df.head(30), use_container_width=True, hide_index=True, height=220)

    if st.button(
        "导入为该客户术语库",
        type="primary",
        use_container_width=True,
        key=f"{key_prefix}_import_btn",
    ):
        try:
            import_stats, import_report = import_customer_glossary(
                user,
                customer_id,
                df,
                conflict_mode,
            )
            combined_report = parse_report + import_report
            combined_stats = {
                **parse_stats,
                **import_stats,
                "error_rows": parse_stats.get("error_rows", 0) + import_stats.get("error_rows", 0),
            }
            st.session_state[f"{key_prefix}_last_import_report"] = combined_report
            st.session_state[f"{key_prefix}_last_import_stats"] = combined_stats
            st.success(
                f"导入完成：新增 active **{import_stats['success_imported']}** 条，"
                f"覆盖 **{import_stats['overwritten']}** 条，"
                f"跳过已有 **{import_stats['skipped_existing']}** 条，"
                f"待审批 **{import_stats['pending_count']}** 条。"
            )
        except Exception as exc:
            st.error(str(exc))

    last_report = st.session_state.get(f"{key_prefix}_last_import_report")
    last_stats = st.session_state.get(f"{key_prefix}_last_import_stats")
    if last_report and last_stats:
        st.caption(
            f"最近一次导入报告：总行数 {last_stats.get('total_rows', 0)}，"
            f"成功导入 {last_stats.get('success_imported', 0)}，"
            f"跳过空行 {last_stats.get('skipped_blank', 0)}，"
            f"重复术语 {last_stats.get('duplicate_terms', 0)}，"
            f"覆盖术语 {last_stats.get('overwritten', 0)}，"
            f"待审批 {last_stats.get('pending_count', 0)}，"
            f"错误行 {last_stats.get('error_rows', 0)}。"
        )
        st.download_button(
            "⬇️ 下载 import_report.xlsx",
            data=build_import_report_xlsx(last_report),
            file_name="import_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"{key_prefix}_import_report_dl",
        )


@st.fragment
def render_excel_jobs_panel(current_user: dict) -> None:
    excel_jobs = list_translation_jobs(current_user["username"], "Excel") if current_user else []
    render_worker_health_indicator()
    if not excel_jobs:
        return

    st.divider()
    st.subheader("Excel 后台任务")
    st.caption("Excel 翻译由独立 worker 处理。清空上传框不会取消后台任务。")
    refresh_col, cancel_col = st.columns(2)
    with refresh_col:
        if st.button("刷新 Excel 任务状态", use_container_width=True, key="refresh_excel_jobs_btn"):
            st.rerun(scope="fragment")
    with cancel_col:
        if st.button("取消等待/翻译中的 Excel", use_container_width=True, key="cancel_running_excel_jobs_btn"):
            cancel_result = cancel_excel_translation_jobs(current_user["username"])
            if cancel_result["cancelled_count"]:
                st.success(f"已取消 {cancel_result['cancelled_count']} 个 Excel 任务。")
            else:
                st.info("当前没有等待中或翻译中的 Excel 任务。")
            st.rerun(scope="fragment")

    complete_jobs = [job for job in excel_jobs if job["status"] == "complete"]
    if len(complete_jobs) > 1:
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w") as zf:
            for job in complete_jobs:
                full_job = get_translation_job(job["job_id"], current_user["username"])
                if not full_job:
                    continue
                base = full_job["source_file_name"].rsplit(".", 1)[0]
                if full_job.get("result_file"):
                    zf.writestr(f"{base}_中文版.xlsx", full_job["result_file"])
                if full_job.get("result_report"):
                    zf.writestr(f"{base}_翻译报告.xlsx", full_job["result_report"])
        st.download_button(
            label=f"⬇️  打包下载全部 {len(complete_jobs)} 个译文（ZIP）",
            data=zip_buf.getvalue(),
            file_name="translated_excels.zip",
            mime="application/zip",
            use_container_width=True,
            type="primary",
            key="excel_jobs_zip_dl",
        )

    for job in excel_jobs:
        label = {
            "queued": "等待中",
            "running": "翻译中",
            "complete": "已完成",
            "failed": "失败",
        }.get(job["status"], job["status"])
        with st.expander(
            f"{label} · {job['source_file_name']} · {job['updated_at']}",
            expanded=job["status"] in {"running", "failed"},
        ):
            st.progress(float(job.get("progress") or 0), text=job.get("message") or label)
            if job.get("error") and job.get("error") != PDF_JOB_CANCELLED_ERROR:
                st.error(job["error"])
            if job.get("error") == PDF_JOB_CANCELLED_ERROR:
                st.info("该任务已取消。可以重新上传文件开始翻译。")
            if job["status"] == "complete":
                full_job = get_translation_job(job["job_id"], current_user["username"])
                if not full_job:
                    continue
                meta = json.loads(full_job.get("result_meta") or "{}")
                review_summary = meta.get("review_summary") or {}
                if review_summary.get("summary_text"):
                    st.info(review_summary["summary_text"])
                st.caption(
                    f"文字单元格 **{meta.get('translated_cell_count', 0)}** 个"
                    f"｜图片译文 TextBox **{meta.get('translated_image_count', 0)}** 条"
                )
                base = full_job["source_file_name"].rsplit(".", 1)[0]
                dl1, dl2 = st.columns(2)
                with dl1:
                    st.download_button(
                        label="⬇️  下载中文 Excel",
                        data=full_job.get("result_file") or b"",
                        file_name=f"{base}_中文版.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"excel_job_dl_{job['job_id']}",
                    )
                with dl2:
                    st.download_button(
                        label="⬇️  下载翻译报告",
                        data=full_job.get("result_report") or b"",
                        file_name=f"{base}_翻译报告.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"excel_report_dl_{job['job_id']}",
                    )


with tab_excel:
    if "excel_upload_reset_nonce" not in st.session_state:
        st.session_state["excel_upload_reset_nonce"] = 0
    if st.button(
        "清空 Excel 翻译并重新上传",
        use_container_width=True,
        key="reset_excel_translation_btn",
    ):
        st.session_state.pop("excel_batch_results", None)
        st.session_state["excel_upload_reset_nonce"] += 1
        st.info("Excel 上传区域已重置，可以重新上传。后台任务不会因此取消；如需停止请点击取消 Excel 任务。")
        st.rerun()

    excel_upload_nonce = st.session_state["excel_upload_reset_nonce"]
    excel_files = st.file_uploader(
        "上传待翻译 Excel（可一次选择多个文件）",
        type=["xlsx"],
        accept_multiple_files=True,
        key=f"excel_uploader_{excel_upload_nonce}",
    )
    selected_glossary_count = len(get_customer_glossary_df(selected_customer_id)) if selected_customer_id else 0
    st.caption(f"📚 当前客户术语库：**{selected_customer_id or '未选择'}** · {selected_glossary_count} 条 active 术语")

    excel_scope_choice = st.radio(
        "翻译范围",
        ["all", "workmanship_auto", "manual"],
        format_func=lambda v: {
            "all": "全部 Sheet",
            "workmanship_auto": "自动识别做工 Sheet",
            "manual": "手动选择 Sheet",
        }[v],
        horizontal=True,
        key="excel_scope_choice",
    )
    excel_scope_configs: dict[str, dict] = {}
    excel_scope_errors: list[str] = []
    if excel_files and excel_scope_choice != "all":
        for file_idx, ef in enumerate(excel_files):
            excel_bytes_for_scope = ef.getvalue()
            detection_rows = []
            try:
                if excel_scope_choice == "workmanship_auto":
                    detection_rows = detect_workmanship_sheets(excel_bytes_for_scope)
                    sheet_names = [row["sheet_name"] for row in detection_rows]
                    default_sheets = [
                        row["sheet_name"] for row in detection_rows
                        if row["is_workmanship"]
                    ]
                else:
                    wb_for_scope = openpyxl.load_workbook(
                        io.BytesIO(excel_bytes_for_scope),
                        read_only=True,
                        data_only=True,
                    )
                    sheet_names = wb_for_scope.sheetnames
                    wb_for_scope.close()
                    default_sheets = sheet_names
                with st.expander(f"Sheet 范围确认：{ef.name}", expanded=True):
                    if detection_rows:
                        preview_df = pd.DataFrame([{
                            "Sheet": row["sheet_name"],
                            "判断": row["verdict"],
                            "分数": row["score"],
                            "原因": row["reason"],
                        } for row in detection_rows])
                        st.dataframe(preview_df, use_container_width=True, hide_index=True, height=240)
                    selected_sheets = st.multiselect(
                        "本文件实际翻译 Sheet",
                        options=sheet_names,
                        default=default_sheets,
                        key=f"excel_scope_sheets_{file_idx}_{ef.name}",
                    )
                if not selected_sheets:
                    excel_scope_errors.append(f"{ef.name}：至少选择 1 个 Sheet。")
                excel_scope_configs[ef.name] = {
                    "selected_sheets": selected_sheets,
                    "detection": detection_rows,
                    "sheet_names": sheet_names,
                }
            except Exception as exc:
                excel_scope_errors.append(f"{ef.name}：范围识别失败：{exc}")
        for err in excel_scope_errors:
            st.error(err)

    customer_ready = bool(selected_customer_id and can_use_customer_glossary(current_user, selected_customer_id))
    can_start_excel = bool(excel_files and api_key and customer_ready and not excel_scope_errors)
    if not can_start_excel:
        missing = []
        if not excel_files:
            missing.append("Excel 文件")
        if not api_key:
            missing.append("OpenAI API Key")
        if not customer_ready:
            missing.append("有权限的客户")
        st.info(f"请先提供：{'、'.join(missing)}")

    translate_images_checked = st.checkbox(
        "🧪 同时翻译图片内文字（实验性：生成可编辑 TextBox，不改图片像素）",
        key="excel_translate_images_checkbox",
    )
    st.caption("勾选后，图片里识别到的英文标注会生成黄底中文 TextBox 叠在图片对应位置，"
               "可在 Excel/WPS 里直接双击编辑、拖动。坐标是视觉模型估算的，密集排版的图"
               "可能不够精确，请翻译后打开核对。不勾选则图片完全原样保留。")

    start_excel_btn = st.button(
        "🚀  开始翻译 Excel",
        disabled=not can_start_excel,
        use_container_width=True,
        type="primary",
        key="start_excel_btn",
    )

    if start_excel_btn:
        created_jobs = []
        for ef in excel_files:
            scope_cfg = excel_scope_configs.get(ef.name, {})
            selected_sheets = None if excel_scope_choice == "all" else scope_cfg.get("selected_sheets")
            scope_detection = scope_cfg.get("detection") if excel_scope_choice != "all" else []
            job_id = create_external_excel_translation_job(
                username=current_user["username"],
                customer_id=selected_customer_id,
                source_file_name=ef.name,
                input_bytes=ef.getvalue(),
                config={
                    "scope_mode": excel_scope_choice,
                    "selected_sheets": selected_sheets,
                    "scope_detection": scope_detection,
                    "translate_images": translate_images_checked,
                }
            )
            created_jobs.append(job_id)
        st.session_state["last_excel_job_ids"] = created_jobs
        st.success(f"已创建 **{len(created_jobs)}** 个 Excel 后台翻译任务。可以稍后回到本页下载结果。")
        st.info("任务已进入等待队列，将由独立 worker 执行。")

    render_excel_jobs_panel(current_user)

# ════════════════════════════════════════════════════════════════════════════
#  Customer Glossary Tab
# ════════════════════════════════════════════════════════════════════════════
with tab_glossary:
    if not selected_customer_id or not can_view_customer_glossary(current_user, selected_customer_id):
        st.error("请先选择一个有权限访问的客户。")
    else:
        customer = get_customer(selected_customer_id)
        st.caption(
            f"当前客户：**{customer['customer_code']} / {customer['customer_name']}** ｜ "
            f"小组：**{customer['group_name']}** ｜ 负责人：**{customer['assigned_staff_username']}**"
        )

        terms = get_active_glossary_terms(selected_customer_id)
        terms_df = pd.DataFrame(terms)
        if terms_df.empty:
            st.info("当前客户还没有 active 术语。")
        else:
            display_df = terms_df[
                ["english_term", "chinese_translation", "normalized_key", "note", "updated_by", "updated_at"]
            ].rename(columns={
                "english_term": "英文术语",
                "chinese_translation": "中文翻译",
                "normalized_key": "normalized_key",
                "note": "备注",
                "updated_by": "最后更新人",
                "updated_at": "最后更新时间",
            })
            st.dataframe(display_df, use_container_width=True, height=260, hide_index=True)

        st.divider()
        st.subheader("提交术语修改申请")
        st.caption("新增、修改、删除只生成 pending 申请；正式术语库不会被直接改动。")

        can_submit = can_submit_glossary_change(current_user, selected_customer_id)
        if not can_submit:
            st.warning("当前用户无权提交该客户的术语修改申请。")
        else:
            add_tab, update_tab, delete_tab = st.tabs(["申请新增术语", "申请修改术语", "申请删除术语"])

            with add_tab:
                with st.form("glossary_add_request_form"):
                    en_new = st.text_input("英文术语", key="add_en")
                    zh_new = st.text_input("中文翻译", key="add_zh")
                    note = st.text_area("备注", key="add_note")
                    submitted = st.form_submit_button("提交新增申请", type="primary", use_container_width=True)
                if submitted:
                    if not en_new.strip() or not zh_new.strip():
                        st.error("英文术语和中文翻译不能为空。")
                    else:
                        req_id = create_glossary_change_request(
                            current_user,
                            selected_customer_id,
                            "add",
                            english_term_new=en_new,
                            chinese_translation_new=zh_new,
                            note=note,
                        )
                        st.success(f"已提交新增申请，申请编号 #{req_id}。等待公司管理员审批。")

            with update_tab:
                if not terms:
                    st.info("当前客户没有可修改的 active 术语。")
                else:
                    term_options = [t["english_term"] for t in terms]
                    selected_term = st.selectbox("选择要修改的术语", term_options, key="update_term_select")
                    old_term = get_active_glossary_term(selected_customer_id, selected_term)
                    with st.form("glossary_update_request_form"):
                        st.text_input("原英文术语", value=old_term["english_term"], disabled=True)
                        st.text_input("原中文翻译", value=old_term["chinese_translation"], disabled=True)
                        en_new = st.text_input("新英文术语", value=old_term["english_term"], key="update_en")
                        zh_new = st.text_input("新中文翻译", value=old_term["chinese_translation"], key="update_zh")
                        note = st.text_area("备注", value=old_term.get("note") or "", key="update_note")
                        submitted = st.form_submit_button("提交修改申请", type="primary", use_container_width=True)
                    if submitted:
                        if not en_new.strip() or not zh_new.strip():
                            st.error("新英文术语和新中文翻译不能为空。")
                        else:
                            req_id = create_glossary_change_request(
                                current_user,
                                selected_customer_id,
                                "update",
                                english_term_old=old_term["english_term"],
                                chinese_translation_old=old_term["chinese_translation"],
                                english_term_new=en_new,
                                chinese_translation_new=zh_new,
                                note=note,
                            )
                            st.success(f"已提交修改申请，申请编号 #{req_id}。等待公司管理员审批。")

            with delete_tab:
                if not terms:
                    st.info("当前客户没有可删除的 active 术语。")
                else:
                    term_options = [t["english_term"] for t in terms]
                    selected_term = st.selectbox("选择要删除的术语", term_options, key="delete_term_select")
                    old_term = get_active_glossary_term(selected_customer_id, selected_term)
                    with st.form("glossary_delete_request_form"):
                        st.text_input("英文术语", value=old_term["english_term"], disabled=True)
                        st.text_input("中文翻译", value=old_term["chinese_translation"], disabled=True)
                        note = st.text_area("删除原因 / 备注", key="delete_note")
                        submitted = st.form_submit_button("提交删除申请", type="primary", use_container_width=True)
                    if submitted:
                        req_id = create_glossary_change_request(
                            current_user,
                            selected_customer_id,
                            "delete",
                            english_term_old=old_term["english_term"],
                            chinese_translation_old=old_term["chinese_translation"],
                            note=note,
                        )
                        st.success(f"已提交删除申请，申请编号 #{req_id}。等待公司管理员审批。")

        st.divider()
        render_customer_glossary_import_panel(
            current_user,
            selected_customer_id,
            "selected_customer_glossary",
        )


# ════════════════════════════════════════════════════════════════════════════
#  Term Candidates Tab
# ════════════════════════════════════════════════════════════════════════════
with tab_candidates:
    accessible_customers = get_accessible_customers(current_user)
    if not accessible_customers:
        st.warning("当前用户没有可访问客户。")
    elif not selected_customer_id:
        st.info("请先在页面顶部选择客户。")
    else:
        candidate_customer_labels = {
            c["customer_id"]: f"{c['customer_code']} / {c['customer_name']}"
            for c in accessible_customers
        }
        current_customer_label = candidate_customer_labels.get(selected_customer_id, selected_customer_id)
        candidate_customer = selected_customer_id
        show_all_candidates = False
        fc, fs = st.columns(2)
        with fc:
            st.caption(f"当前客户：**{current_customer_label}**")
            if can_approve_glossary_change(current_user):
                show_all_candidates = st.toggle(
                    "查看全部客户候选",
                    value=False,
                    key="candidate_show_all_toggle",
                )
        with fs:
            candidate_status = st.selectbox(
                "状态",
                ["draft", "selected", "submitted", "ignored", "approved", "rejected", "全部"],
                key="candidate_status_filter",
            )

        candidates = list_term_candidates(
            current_user,
            customer_id=None if show_all_candidates else candidate_customer,
            status=candidate_status,
        )
        if not candidates:
            st.info("当前没有符合条件的术语候选。翻译 PDF / Excel 后会自动生成候选。")
        else:
            rows = []
            for c in candidates:
                try:
                    variants = "; ".join(json.loads(c.get("variants") or "[]"))
                except json.JSONDecodeError:
                    variants = c.get("variants") or c["original_term"]
                rows.append({
                    "是否提交": False,
                    "忽略": False,
                    "candidate_id": c["candidate_id"],
                    "customer_id": c["customer_id"],
                    "客户": f"{c.get('customer_code') or c['customer_id']} / {c.get('customer_name') or ''}",
                    "normalized_key": c["normalized_term"],
                    "英文术语": c["original_term"],
                    "所有变体": variants or c["original_term"],
                    "AI建议中文": c["ai_suggested_translation"],
                    "人工确认中文": c["final_translation"] or c["ai_suggested_translation"],
                    "出现次数": c["frequency"],
                    "来源文件": c["source_file_name"],
                    "来源类型": c["source_type"],
                    "来源位置": c["page_or_sheet"],
                    "单元格": c["cell_coordinate"],
                    "上下文": c["context_sentence"],
                    "置信度": c["confidence"],
                    "matched_by": c.get("matched_by") or "no_match",
                    "命中术语": c.get("matched_glossary_term") or "",
                    "冲突提示": c.get("conflict_warning") or "",
                    "状态": c["status"],
                    "备注": "",
                })
            editor_df = pd.DataFrame(rows)
            edited_df = st.data_editor(
                editor_df,
                use_container_width=True,
                hide_index=True,
                height=420,
                disabled=[
                    "candidate_id", "customer_id", "客户", "normalized_key",
                    "英文术语", "所有变体", "AI建议中文",
                    "出现次数", "来源文件", "来源类型", "来源位置", "单元格",
                    "上下文", "置信度", "matched_by", "命中术语", "冲突提示", "状态",
                ],
                column_config={
                    "是否提交": st.column_config.CheckboxColumn("是否提交"),
                    "忽略": st.column_config.CheckboxColumn("忽略"),
                    "人工确认中文": st.column_config.TextColumn("人工确认中文", required=False),
                    "备注": st.column_config.TextColumn("备注"),
                },
                key="term_candidates_editor",
            )

            submit_col, ignore_col = st.columns(2)
            with submit_col:
                submit_candidates = st.button(
                    "一键提交审核",
                    type="primary",
                    use_container_width=True,
                    key="submit_term_candidates_btn",
                )
            with ignore_col:
                ignore_candidates = st.button(
                    "忽略勾选候选",
                    use_container_width=True,
                    key="ignore_term_candidates_btn",
                )

            if submit_candidates:
                selected_rows = []
                for _, row in edited_df[edited_df["是否提交"] == True].iterrows():
                    selected_rows.append({
                        "candidate_id": int(row["candidate_id"]),
                        "customer_id": row["customer_id"],
                        "final_translation": row["人工确认中文"],
                        "note": row.get("备注", ""),
                    })
                if not selected_rows:
                    st.warning("请先勾选要提交的候选术语。")
                else:
                    submitted_count, errors = submit_term_candidates_for_approval(current_user, selected_rows)
                    if submitted_count:
                        st.success(f"已提交 **{submitted_count}** 条术语候选，等待管理员审批。")
                    if errors:
                        st.warning("；".join(errors[:10]))
                    st.rerun()

            if ignore_candidates:
                ignore_ids = [
                    int(row["candidate_id"])
                    for _, row in edited_df[edited_df["忽略"] == True].iterrows()
                ]
                if not ignore_ids:
                    st.warning("请先勾选要忽略的候选术语。")
                else:
                    placeholders = ",".join("?" for _ in ignore_ids)
                    with get_db_connection() as conn:
                        conn.execute(
                            f"""
                            UPDATE term_candidates
                            SET status = 'ignored'
                            WHERE candidate_id IN ({placeholders})
                            """,
                            ignore_ids,
                        )
                    st.success(f"已忽略 **{len(ignore_ids)}** 条候选。")
                    st.rerun()


# ════════════════════════════════════════════════════════════════════════════
#  Approval Tab
# ════════════════════════════════════════════════════════════════════════════
if tab_approval is not None:
    with tab_approval:
        st.caption("仅 company_admin 可见。审批通过后才会写入正式客户术语库。")

        all_customers = get_accessible_customers(current_user)
        customer_filter_options = ["全部"] + [c["customer_id"] for c in all_customers]
        customer_labels = {
            c["customer_id"]: f"{c['customer_code']} / {c['customer_name']}"
            for c in all_customers
        }
        fc, fa, fs = st.columns(3)
        with fc:
            customer_filter = st.selectbox(
                "客户筛选",
                customer_filter_options,
                format_func=lambda cid: "全部" if cid == "全部" else customer_labels.get(cid, cid),
                key="approval_customer_filter",
            )
        with fa:
            action_filter = st.selectbox("申请类型", ["全部", "add", "update", "delete"], key="approval_action_filter")
        with fs:
            submitted_by_filter = st.text_input("申请人筛选", key="approval_submitter_filter")

        pending = list_glossary_change_requests(
            status="pending",
            customer_id=None if customer_filter == "全部" else customer_filter,
            submitted_by=submitted_by_filter.strip() or None,
            action_type=None if action_filter == "全部" else action_filter,
        )

        if not pending:
            st.info("当前没有符合条件的 pending 申请。")
        else:
            pending_df = pd.DataFrame(pending)
            st.dataframe(
                pending_df[[
                    "request_id", "candidate_id", "customer_code", "customer_name", "action_type",
                    "english_term_old", "chinese_translation_old",
                    "english_term_new", "chinese_translation_new",
                    "submitted_by", "submitted_at",
                ]],
                use_container_width=True,
                hide_index=True,
                height=260,
            )

            request_ids = [int(r["request_id"]) for r in pending]
            selected_request_id = st.selectbox("选择要审批的申请", request_ids, key="approval_request_select")
            selected_request = next(r for r in pending if int(r["request_id"]) == int(selected_request_id))

            st.markdown("#### 申请详情")
            detail_cols = st.columns(2)
            with detail_cols[0]:
                st.write(f"客户：**{selected_request['customer_code']} / {selected_request['customer_name']}**")
                st.write(f"类型：`{selected_request['action_type']}`")
                if selected_request.get("candidate_id"):
                    st.write(f"候选编号：`{selected_request['candidate_id']}`")
                st.write(f"申请人：**{selected_request['submitted_by']}**")
                st.write(f"提交时间：{selected_request['submitted_at']}")
            with detail_cols[1]:
                st.write(f"旧英文：{selected_request['english_term_old'] or ''}")
                st.write(f"旧中文：{selected_request['chinese_translation_old'] or ''}")
                st.write(f"新英文：{selected_request['english_term_new'] or ''}")
                st.write(f"新中文：{selected_request['chinese_translation_new'] or ''}")
            if selected_request.get("note"):
                st.info(f"备注：{selected_request['note']}")

            with st.form("approval_decision_form"):
                review_comment = st.text_area("审核意见", key="approval_comment")
                approve_col, reject_col = st.columns(2)
                with approve_col:
                    approve_clicked = st.form_submit_button("通过申请", type="primary", use_container_width=True)
                with reject_col:
                    reject_clicked = st.form_submit_button("拒绝申请", use_container_width=True)

            if approve_clicked:
                try:
                    approve_glossary_change_request(selected_request_id, current_user, review_comment)
                    st.success(f"已通过申请 #{selected_request_id}，正式术语库已更新。")
                    st.rerun()
                except Exception as exc:
                    st.error(str(exc))

            if reject_clicked:
                try:
                    reject_glossary_change_request(selected_request_id, current_user, review_comment)
                    st.success(f"已拒绝申请 #{selected_request_id}，正式术语库未改变。")
                    st.rerun()
                except Exception as exc:
                    st.error(str(exc))


# ════════════════════════════════════════════════════════════════════════════
#  User Admin Tab
# ════════════════════════════════════════════════════════════════════════════
if tab_user_admin is not None:
    with tab_user_admin:
        st.caption("仅 company_admin 可见。客户分配以「客户管理」中的负责人为准。")
        users = list_users()
        users_df = pd.DataFrame(users)
        if not users_df.empty:
            display_users = users_df.copy()
            display_users["assigned_customer_ids"] = display_users["assigned_customer_ids"].apply(
                lambda ids: ", ".join(ids) if ids else ""
            )
            st.dataframe(
                display_users[["username", "role", "group_name", "assigned_customer_ids"]],
                use_container_width=True,
                hide_index=True,
                height=240,
            )

        st.divider()
        add_user_tab, edit_user_tab = st.tabs(["新增账号", "修改账号"])

        with add_user_tab:
            with st.form("create_user_form"):
                username = st.text_input("用户名", key="create_user_username")
                password = st.text_input("初始密码", type="password", key="create_user_password")
                role = st.selectbox(
                    "角色",
                    ["staff", "group_leader", "company_admin"],
                    key="create_user_role",
                )
                group_name = st.text_input("小组", key="create_user_group")
                submitted = st.form_submit_button("创建账号", type="primary", use_container_width=True)
            if submitted:
                try:
                    create_user(username, password, role, group_name)
                    st.success(f"已创建账号 {username.strip()}。")
                    st.rerun()
                except Exception as exc:
                    st.error(str(exc))

        with edit_user_tab:
            if not users:
                st.info("当前没有用户。")
            else:
                usernames = [u["username"] for u in users]
                selected_username = st.selectbox("选择账号", usernames, key="edit_user_select")
                selected_user = next(u for u in users if u["username"] == selected_username)
                with st.form("update_user_form"):
                    role = st.selectbox(
                        "角色",
                        ["staff", "group_leader", "company_admin"],
                        index=["staff", "group_leader", "company_admin"].index(selected_user["role"]),
                        key="edit_user_role",
                    )
                    group_name = st.text_input(
                        "小组",
                        value=selected_user.get("group_name") or "",
                        key="edit_user_group",
                    )
                    new_password = st.text_input(
                        "新密码（留空则不修改）",
                        type="password",
                        key="edit_user_password",
                    )
                    st.caption(
                        "该账号负责客户："
                        + (", ".join(selected_user.get("assigned_customer_ids") or []) or "无")
                    )
                    submitted = st.form_submit_button("保存账号修改", type="primary", use_container_width=True)
                if submitted:
                    try:
                        update_user(selected_username, role, group_name, new_password)
                        if selected_username == current_user["username"]:
                            st.session_state["current_user"] = authenticate_user(
                                selected_username,
                                new_password if new_password.strip() else "",
                            ) or {**current_user, "role": role, "group_name": group_name}
                        st.success(f"已更新账号 {selected_username}。")
                        st.rerun()
                    except Exception as exc:
                        st.error(str(exc))


# ════════════════════════════════════════════════════════════════════════════
#  Customer Admin Tab
# ════════════════════════════════════════════════════════════════════════════
if tab_customer_admin is not None:
    with tab_customer_admin:
        st.caption("仅 company_admin 可见。创建/修改负责人后会同步 staff 的 assigned_customer_ids。")
        customers = list_customers_with_term_counts(current_user)
        staff_usernames = get_staff_usernames()
        customers_df = pd.DataFrame(customers)
        if not customers_df.empty:
            display_customers = customers_df[[
                "customer_id", "customer_code", "customer_name",
                "group_name", "assigned_staff_username", "term_count",
                "created_at", "note",
            ]].rename(columns={
                "customer_id": "客户ID",
                "customer_code": "客户代码",
                "customer_name": "客户名称",
                "group_name": "所属小组",
                "assigned_staff_username": "负责职员",
                "term_count": "当前术语数量",
                "created_at": "创建时间",
                "note": "备注",
            })
            st.dataframe(display_customers, use_container_width=True, hide_index=True, height=260)
        else:
            st.info("当前没有客户。")

        st.divider()
        add_customer_tab, import_customer_tab, edit_customer_tab = st.tabs([
            "创建新客户",
            "上传/更新术语库",
            "修改客户",
        ])

        with add_customer_tab:
            if not staff_usernames:
                st.warning("请先创建 staff 账号，再新增客户。")
            else:
                with st.form("create_customer_form"):
                    customer_name = st.text_input("客户名称", key="create_customer_name")
                    customer_code = st.text_input("客户代码", placeholder="例如 EL 或 CUST004", key="create_customer_code")
                    group_name = st.text_input("小组", key="create_customer_group")
                    assigned_staff_username = st.selectbox(
                        "负责业务员 / 普通职员",
                        staff_usernames,
                        key="create_customer_staff",
                    )
                    note = st.text_area("备注（可选）", key="create_customer_note")
                    submitted = st.form_submit_button("创建客户", type="primary", use_container_width=True)
                if submitted:
                    try:
                        new_customer_id = create_customer_auto_id(
                            customer_name,
                            customer_code,
                            group_name,
                            assigned_staff_username,
                            note,
                        )
                        st.session_state["_pending_selected_customer_id"] = new_customer_id
                        st.success(f"已创建客户 {customer_code.strip()}，可立即在 PDF / Excel 翻译中选择。")
                        st.rerun()
                    except Exception as exc:
                        st.error(str(exc))

        with import_customer_tab:
            if not customers:
                st.info("请先创建客户。")
            else:
                customer_ids = [c["customer_id"] for c in customers]
                labels = {
                    c["customer_id"]: f"{c['customer_code']} / {c['customer_name']} / {c['group_name']}"
                    for c in customers
                }
                default_customer = st.session_state.get("selected_customer_id")
                default_index = customer_ids.index(default_customer) if default_customer in customer_ids else 0
                import_customer_id = st.selectbox(
                    "选择客户",
                    customer_ids,
                    index=default_index,
                    format_func=lambda cid: labels.get(cid, cid),
                    key="admin_import_customer_select",
                )
                selected_terms = get_customer_glossary_df(import_customer_id)
                st.caption(f"当前 active 术语：**{len(selected_terms)}** 条")
                dl_col, view_col = st.columns(2)
                with dl_col:
                    st.download_button(
                        "⬇️ 下载该客户术语库",
                        data=glossary_df_to_xlsx_bytes(selected_terms),
                        file_name=f"{labels.get(import_customer_id, import_customer_id).split(' / ')[0]}_glossary.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="admin_customer_glossary_dl",
                    )
                with view_col:
                    if st.button("查看术语库", use_container_width=True, key="admin_view_customer_terms"):
                        st.session_state["admin_show_customer_terms"] = import_customer_id
                if st.session_state.get("admin_show_customer_terms") == import_customer_id:
                    st.dataframe(selected_terms, use_container_width=True, hide_index=True, height=220)

                st.divider()
                render_customer_glossary_import_panel(
                    current_user,
                    import_customer_id,
                    "admin_customer_glossary",
                )

        with edit_customer_tab:
            if not customers:
                st.info("当前没有客户。")
            elif not staff_usernames:
                st.warning("请先创建 staff 账号。")
            else:
                customer_ids = [c["customer_id"] for c in customers]
                selected_customer = st.selectbox("选择客户", customer_ids, key="edit_customer_select")
                customer = next(c for c in customers if c["customer_id"] == selected_customer)
                default_staff_index = (
                    staff_usernames.index(customer["assigned_staff_username"])
                    if customer["assigned_staff_username"] in staff_usernames
                    else 0
                )
                with st.form("update_customer_form"):
                    st.text_input("客户 ID", value=customer["customer_id"], disabled=True)
                    customer_code = st.text_input(
                        "客户代码",
                        value=customer["customer_code"],
                        key="edit_customer_code",
                    )
                    customer_name = st.text_input(
                        "客户名称",
                        value=customer["customer_name"],
                        key="edit_customer_name",
                    )
                    group_name = st.text_input(
                        "小组",
                        value=customer["group_name"],
                        key="edit_customer_group",
                    )
                    assigned_staff_username = st.selectbox(
                        "负责人",
                        staff_usernames,
                        index=default_staff_index,
                        key="edit_customer_staff",
                    )
                    note = st.text_area(
                        "备注",
                        value=customer.get("note") or "",
                        key="edit_customer_note",
                    )
                    submitted = st.form_submit_button("保存客户修改", type="primary", use_container_width=True)
                if submitted:
                    try:
                        update_customer(
                            selected_customer,
                            customer_name,
                            customer_code,
                            group_name,
                            assigned_staff_username,
                        )
                        with get_db_connection() as conn:
                            conn.execute(
                                "UPDATE customers SET note = ? WHERE customer_id = ?",
                                (note.strip(), selected_customer),
                            )
                        st.success(f"已更新客户 {selected_customer}。")
                        st.rerun()
                    except Exception as exc:
                        st.error(str(exc))
