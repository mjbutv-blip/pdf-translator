import io
import json
import os
import re
import tempfile

import anthropic
import fitz
import openpyxl
import streamlit as st
from pathlib import Path

DEFAULT_FONT = Path(__file__).parent / "font.ttf"


# ── Glossary helpers ───────────────────────────────────────────────────────────

def load_glossary(data: bytes) -> dict[str, str]:
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    chi_idx = eng_idx = None
    for i, h in enumerate(headers):
        if "中文" in h or h.lower() == "chinese":
            chi_idx = i
        if "英文" in h or h.lower() == "english":
            eng_idx = i
    if chi_idx is None or eng_idx is None:
        chi_idx, eng_idx = 0, 1
    g = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        eng = row[eng_idx] if eng_idx < len(row) else None
        chi = row[chi_idx] if chi_idx < len(row) else None
        if eng and chi:
            g[str(eng).strip()] = str(chi).strip()
    return g


def relevant_glossary(text: str, glossary: dict) -> dict:
    tl = text.lower()
    return {k: v for k, v in glossary.items() if k.lower() in tl}


# ── PDF translation ────────────────────────────────────────────────────────────

def translate_block(client: anthropic.Anthropic, text: str, glossary: dict) -> dict:
    rel = relevant_glossary(text, glossary)
    gloss_block = ""
    if rel:
        lines = "\n".join(f"  {k} → {v}" for k, v in rel.items())
        gloss_block = f"强制术语对照（务必照搬，不得自行发挥）：\n{lines}\n\n"

    prompt = (
        "你是一名专业服装行业翻译，请将以下英文文本翻译成地道的中文服装术语。\n\n"
        f"{gloss_block}"
        f"待翻译文本：\n{text}\n\n"
        "规则：\n"
        "1. 凡出现上方术语对照中的词汇，必须使用对照表中的中文译法，禁止替换。\n"
        "2. 保留款号、数字、尺码、货号等编码不翻译。\n"
        "3. 识别文本中出现的、术语对照表里**未收录**的服装/纺织行业专业英文词汇，放入 unrecorded_terms。\n"
        "4. 只返回 JSON，不要有任何多余的文字或 markdown 代码块。\n\n"
        '返回格式：{"translated_text": "中文结果", "unrecorded_terms": ["term1", "term2"]}'
    )
    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=2048,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = msg.content[0].text.strip()
    raw = re.sub(r"^```[a-z]*\n?", "", raw)
    raw = re.sub(r"\n?```$", "", raw)
    m = re.search(r"\{.*\}", raw, re.DOTALL)
    if m:
        raw = m.group()
    return json.loads(raw)


def _insert_text(page, bbox, text: str, size: float, font_path: str):
    rect = fitz.Rect(bbox)
    for s in [size, size * 0.85, size * 0.7, 8.0, 7.0]:
        if page.insert_textbox(rect, text, fontname="myfont", fontfile=font_path,
                               fontsize=s, color=(0, 0, 0), align=0) >= 0:
            return
    page.insert_textbox(rect, text, fontname="myfont", fontfile=font_path,
                        fontsize=7.0, color=(0, 0, 0), align=0)


def run_pdf_translation(
    pdf_bytes: bytes,
    glossary_bytes: bytes,
    font_path: str,
    api_key: str,
    on_page,
    on_block,
    on_progress,
) -> tuple[bytes, bytes, int]:
    """Returns (pdf_out, xlsx_out, n_unrecorded_terms)."""
    glossary = load_glossary(glossary_bytes)
    client = anthropic.Anthropic(api_key=api_key)
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    total_pages = len(doc)
    all_unrecorded: set[str] = set()

    page_blocks: list[list[dict]] = []
    for pn in range(total_pages):
        page = doc[pn]
        tb = []
        for blk in page.get_text("dict")["blocks"]:
            if blk["type"] != 0:
                continue
            parts, sizes = [], []
            for ln in blk.get("lines", []):
                for sp in ln.get("spans", []):
                    if sp["text"].strip():
                        parts.append(sp["text"])
                        sizes.append(sp["size"])
                parts.append("\n")
            text = "".join(parts).strip()
            if not text:
                continue
            tb.append({
                "bbox": blk["bbox"],
                "text": text,
                "size": sum(sizes) / len(sizes) if sizes else 10.0,
            })
        page_blocks.append(tb)

    total_blocks = max(sum(len(b) for b in page_blocks), 1)
    processed = 0

    for pn, text_blocks in enumerate(page_blocks):
        page = doc[pn]
        on_page(pn, total_pages, len(text_blocks))

        results = []
        for blk in text_blocks:
            on_block(blk["text"][:60].replace("\n", " "))
            try:
                res = translate_block(client, blk["text"], glossary)
                translated = res.get("translated_text", blk["text"])
                terms = [t for t in res.get("unrecorded_terms", []) if t.strip()]
                all_unrecorded.update(terms)
            except Exception:
                translated = blk["text"]
            results.append({**blk, "translated": translated})
            processed += 1
            on_progress(processed / total_blocks)

        for r in results:
            page.add_redact_annot(fitz.Rect(r["bbox"]), fill=(1, 1, 1))
        page.apply_redactions()
        for r in results:
            _insert_text(page, r["bbox"], r["translated"], r["size"], font_path)

    pdf_buf = io.BytesIO()
    doc.save(pdf_buf, garbage=4, deflate=True)
    doc.close()

    xlsx_buf = io.BytesIO()
    if all_unrecorded:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "未收录术语"
        ws.append(["未收录英文术语"])
        ws["A1"].font = openpyxl.styles.Font(bold=True)
        for term in sorted(all_unrecorded):
            ws.append([term])
        ws.column_dimensions["A"].width = 45
        wb.save(xlsx_buf)

    return pdf_buf.getvalue(), xlsx_buf.getvalue(), len(all_unrecorded)


# ── Excel translation ──────────────────────────────────────────────────────────

def _is_translatable(val) -> bool:
    if not isinstance(val, str):
        return False
    s = val.strip()
    if not s:
        return False
    if s.startswith("="):
        return False
    return bool(re.search(r"[a-zA-Z]", s))


def translate_cell_text(client: anthropic.Anthropic, text: str, glossary: dict) -> str:
    rel = relevant_glossary(text, glossary)
    gloss_block = ""
    if rel:
        lines = "\n".join(f"  {k} → {v}" for k, v in rel.items())
        gloss_block = f"强制术语对照（务必照搬）：\n{lines}\n\n"

    prompt = (
        "你是一名专业服装行业翻译，将下方英文翻译成中文服装术语。\n\n"
        f"{gloss_block}"
        f"待翻译文本：{text}\n\n"
        "规则：\n"
        "1. 对照表中的词汇必须使用对照表给出的中文译法。\n"
        "2. 款号、货号、数字、尺码保持原样，不翻译。\n"
        "3. 只返回翻译结果，不要任何解释或多余文字。"
    )
    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=512,
        messages=[{"role": "user", "content": prompt}],
    )
    return msg.content[0].text.strip()


def run_excel_translation(
    xlsx_bytes: bytes,
    glossary_bytes: bytes,
    api_key: str,
    on_cell,
    on_progress,
) -> tuple[bytes, int]:
    """Translate all text cells in-place; returns (xlsx_out, n_cells_translated)."""
    glossary = load_glossary(glossary_bytes)
    client = anthropic.Anthropic(api_key=api_key)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active

    to_translate = [
        cell
        for row in ws.iter_rows()
        for cell in row
        if _is_translatable(cell.value)
    ]
    total = max(len(to_translate), 1)

    for i, cell in enumerate(to_translate):
        on_cell(str(cell.value)[:60].replace("\n", " "))
        try:
            cell.value = translate_cell_text(client, str(cell.value), glossary)
        except Exception:
            pass  # keep original on failure
        on_progress((i + 1) / total)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), len(to_translate)


# ── Streamlit page ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="服装行业翻译引擎",
    page_icon="🧵",
    layout="centered",
    initial_sidebar_state="collapsed",
)

st.title("🧵 服装行业翻译引擎")
st.caption("支持 PDF 与 Excel (.xlsx) 双格式 · 上传文件 + 术语库 · 调用 Claude 自动翻译为中文")
st.divider()

# ── 文件上传区 ─────────────────────────────────────────────────────────────────
col1, col2 = st.columns(2)
with col1:
    source_file = st.file_uploader(
        "📄 上传待翻译文件（PDF 或 Excel）",
        type=["pdf", "xlsx"],
    )
with col2:
    glossary_file = st.file_uploader("📊 上传术语库 Excel", type=["xlsx", "xls"])

# 仅 PDF 模式需要字体
source_is_pdf = source_file is not None and source_file.name.lower().endswith(".pdf")
source_is_xlsx = source_file is not None and source_file.name.lower().endswith(".xlsx")

if source_is_pdf:
    font_label = (
        f"🔤 上传中文字体 TTF（可选，已检测到默认字体 {DEFAULT_FONT.name}）"
        if DEFAULT_FONT.exists()
        else "🔤 上传中文字体 TTF（必填，未检测到默认字体）"
    )
    font_file = st.file_uploader(font_label, type=["ttf"])
else:
    font_file = None

st.divider()

# ── API Key ────────────────────────────────────────────────────────────────────
api_key = st.text_input(
    "🔑 Anthropic API Key",
    type="password",
    value=os.environ.get("ANTHROPIC_API_KEY", ""),
    placeholder="sk-ant-api03-...",
)
st.divider()

# ── 状态提示 & 开始按钮 ────────────────────────────────────────────────────────
font_ready = bool(font_file or DEFAULT_FONT.exists())

if source_is_pdf:
    can_start = bool(source_file and glossary_file and api_key and font_ready)
elif source_is_xlsx:
    can_start = bool(source_file and glossary_file and api_key)
else:
    can_start = False

if not can_start:
    missing = []
    if not source_file:
        missing.append("待翻译文件（PDF / Excel）")
    if not glossary_file:
        missing.append("术语库 Excel")
    if not api_key:
        missing.append("Anthropic API Key")
    if source_is_pdf and not font_ready:
        missing.append("中文字体 TTF")
    st.info(f"请先提供：{'、'.join(missing)}" if missing else "")

start_btn = st.button(
    "🚀  开始翻译",
    disabled=not can_start,
    use_container_width=True,
    type="primary",
)

# ── 翻译流程 ───────────────────────────────────────────────────────────────────
if start_btn:
    for key in ("pdf_result", "xlsx_result", "excel_result", "n_terms", "n_cells", "mode"):
        st.session_state.pop(key, None)

    progress_bar = st.progress(0.0, text="初始化…")

    # ── PDF 分支 ──
    if source_is_pdf:
        with tempfile.TemporaryDirectory() as tmpdir:
            if font_file:
                fp = os.path.join(tmpdir, "font.ttf")
                with open(fp, "wb") as f:
                    f.write(font_file.read())
                font_path = fp
            else:
                font_path = str(DEFAULT_FONT)

            with st.status("翻译进行中，请稍候…", expanded=True) as status:
                block_ph = st.empty()

                def on_page(pn, total, n_blocks):
                    st.write(f"**── 第 {pn + 1} / {total} 页**　（{n_blocks} 个文本块）")

                def on_block(preview):
                    block_ph.caption(f"▶ 正在翻译：{preview}…")

                def on_progress(frac):
                    progress_bar.progress(frac, text=f"翻译进度 {frac:.0%}")

                try:
                    pdf_out, xlsx_out, n_terms = run_pdf_translation(
                        pdf_bytes=source_file.read(),
                        glossary_bytes=glossary_file.read(),
                        font_path=font_path,
                        api_key=api_key,
                        on_page=on_page,
                        on_block=on_block,
                        on_progress=on_progress,
                    )
                    block_ph.empty()
                    status.update(label="✅ 翻译完成！", state="complete")
                    progress_bar.progress(1.0, text="完成 ✓")
                    st.session_state["mode"]        = "pdf"
                    st.session_state["pdf_result"]  = pdf_out
                    st.session_state["xlsx_result"] = xlsx_out
                    st.session_state["n_terms"]     = n_terms

                except Exception as e:
                    block_ph.empty()
                    status.update(label=f"❌ 出错：{e}", state="error")
                    st.error(str(e))

    # ── Excel 分支 ──
    elif source_is_xlsx:
        with st.status("翻译进行中，请稍候…", expanded=True) as status:
            cell_ph = st.empty()

            def on_cell(preview):
                cell_ph.caption(f"▶ 正在翻译：{preview}…")

            def on_progress(frac):
                progress_bar.progress(frac, text=f"翻译进度 {frac:.0%}")

            try:
                excel_out, n_cells = run_excel_translation(
                    xlsx_bytes=source_file.read(),
                    glossary_bytes=glossary_file.read(),
                    api_key=api_key,
                    on_cell=on_cell,
                    on_progress=on_progress,
                )
                cell_ph.empty()
                status.update(label="✅ 翻译完成！", state="complete")
                progress_bar.progress(1.0, text="完成 ✓")
                st.session_state["mode"]         = "excel"
                st.session_state["excel_result"] = excel_out
                st.session_state["n_cells"]      = n_cells
                st.session_state["source_name"]  = source_file.name

            except Exception as e:
                cell_ph.empty()
                status.update(label=f"❌ 出错：{e}", state="error")
                st.error(str(e))

# ── 下载区 ─────────────────────────────────────────────────────────────────────
mode = st.session_state.get("mode")

if mode == "pdf" and st.session_state.get("pdf_result"):
    st.divider()
    n = st.session_state["n_terms"]
    st.success(f"翻译完成！共识别 **{n}** 条未收录术语。")

    dl1, dl2 = st.columns(2)
    with dl1:
        st.download_button(
            label="⬇️  下载中文 PDF",
            data=st.session_state["pdf_result"],
            file_name="translated_CN.pdf",
            mime="application/pdf",
            use_container_width=True,
            type="primary",
        )
    with dl2:
        if n > 0:
            st.download_button(
                label="⬇️  下载未收录术语 Excel",
                data=st.session_state["xlsx_result"],
                file_name="unrecorded_terms.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.info("本次翻译无未收录术语")

elif mode == "excel" and st.session_state.get("excel_result"):
    st.divider()
    n = st.session_state["n_cells"]
    st.success(f"翻译完成！共翻译 **{n}** 个单元格。")

    src_name = st.session_state.get("source_name", "translated.xlsx")
    out_name = src_name.replace(".xlsx", "_中文版.xlsx")
    st.download_button(
        label="⬇️  下载中文 Excel",
        data=st.session_state["excel_result"],
        file_name=out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )
