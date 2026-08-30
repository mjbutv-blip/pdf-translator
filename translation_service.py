from __future__ import annotations

import io
import json
import traceback
from pathlib import Path
from typing import Any
import zipfile

import openpyxl
import pandas as pd
import fitz

import rfq_identity


from ai_client import error_metadata
from config import DEFAULT_FONT
from db import get_db_connection
from translation_core import *
from translation_core import (
    _load_default_glossary_df,
    _pending_term_keys,
    _read_excel_sheet_rows,
    _user_from_row,
)
from translation_jobs import (
    _pdf_scope_report_rows,
    create_translation_job,
    delete_translation_job as _legacy_delete_translation_job,
    list_candidate_occurrences as _jobs_list_candidate_occurrences,
    list_term_candidates_for_job as _jobs_list_term_candidates_for_job,
    update_translation_job,
)


_core_list_customers = globals()["list_customers"]
_core_get_customer = globals()["get_customer"]
_core_list_term_candidates_for_job = _jobs_list_term_candidates_for_job
_core_list_candidate_occurrences = _jobs_list_candidate_occurrences
_core_list_glossary_change_requests = globals()["list_glossary_change_requests"]
_core_approve_glossary_change_request = globals()["approve_glossary_change_request"]
_core_reject_glossary_change_request = globals()["reject_glossary_change_request"]


def _service_user(username: str | None = None) -> dict[str, Any]:
    return {
        "username": (username or "service").strip() or "service",
        "role": "company_admin",
        "group_name": "",
        "assigned_customer_ids": [],
    }


def _ensure_db_ready() -> None:
    init_db()


def _load_user_for_service(username: str | None) -> dict[str, Any] | None:
    name = str(username or "").strip()
    if not name:
        return None
    with get_db_connection() as conn:
        row = conn.execute(
            "SELECT * FROM users WHERE username = ?",
            (name,),
        ).fetchone()
    return _user_from_row(row) if row else None


def _require_company_admin(username: str | None) -> dict[str, Any]:
    user = _load_user_for_service(username)
    if user is None:
        raise PermissionError("permission denied: reviewer username 不存在")
    if user.get("role") != "company_admin":
        raise PermissionError("permission denied: reviewer 不是 company_admin")
    return user


def _authorize_pdf_customer_unified(
    rfq_access_token: str,
    pdf_customer_id: str,
    *,
    require_write: bool = False,
    require_quote_use: bool = False,
) -> tuple[dict[str, Any], dict[str, Any]]:
    return rfq_identity.authorize_pdf_customer(
        rfq_access_token,
        pdf_customer_id,
        require_write=require_write,
        require_quote_use=require_quote_use,
    )


def _confirmed_pdf_customer_ids_unified(rfq_access_token: str) -> tuple[dict[str, Any], list[str]]:
    return rfq_identity.confirmed_pdf_customer_ids(rfq_access_token)


def _candidate_customer_id(candidate_id: int) -> str | None:
    with get_db_connection() as conn:
        row = conn.execute(
            "SELECT customer_id FROM term_candidates WHERE candidate_id = ?",
            (int(candidate_id),),
        ).fetchone()
    return str(row["customer_id"]) if row else None


def _job_customer_id(job_id: str) -> str | None:
    with get_db_connection() as conn:
        row = conn.execute(
            "SELECT customer_id FROM translation_jobs WHERE job_id = ?",
            (job_id,),
        ).fetchone()
    return str(row["customer_id"]) if row else None


def _glossary_customer_id(glossary_id: int) -> str | None:
    with get_db_connection() as conn:
        row = conn.execute(
            "SELECT customer_id FROM glossary_terms WHERE glossary_id = ?",
            (int(glossary_id),),
        ).fetchone()
    return str(row["customer_id"]) if row else None


def _change_request_customer_id(request_id: int) -> str | None:
    with get_db_connection() as conn:
        row = conn.execute(
            "SELECT customer_id FROM glossary_change_requests WHERE request_id = ?",
            (int(request_id),),
        ).fetchone()
    return str(row["customer_id"]) if row else None


def _authorize_existing_customer_unified(
    rfq_access_token: str,
    customer_id: str | None,
    *,
    require_write: bool = False,
    require_quote_use: bool = False,
) -> tuple[dict[str, Any], dict[str, Any]]:
    if not customer_id:
        raise rfq_identity.UnifiedAuthError("target record does not exist")
    return _authorize_pdf_customer_unified(
        rfq_access_token,
        customer_id,
        require_write=require_write,
        require_quote_use=require_quote_use,
    )


def _json_loads_dict(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    if not value:
        return {}
    try:
        data = json.loads(value)
        return data if isinstance(data, dict) else {}
    except (TypeError, json.JSONDecodeError):
        return {}


def _normalize_result_meta(value: Any) -> dict[str, Any]:
    meta = _json_loads_dict(value)
    if "candidate_ids_reliable" not in meta:
        meta["candidate_ids_reliable"] = False
    if "candidate_ids" not in meta:
        meta["candidate_ids"] = []
    if "candidate_count" not in meta:
        meta["candidate_count"] = None
    if "unrecorded_term_count" not in meta:
        meta["unrecorded_term_count"] = None
    if "review_item_count" not in meta:
        meta["review_item_count"] = None
    return meta


def _job_type_for_file_type(file_type: str) -> str:
    return "PDF" if file_type.lower() == "pdf" else "Excel"


def _safe_filename(filename: str, source_file_name: str | None = None) -> str:
    return (source_file_name or filename or "uploaded_file").strip() or "uploaded_file"


def _font_job_payload(font_path: str | None, resolved_font_path: str | None = None) -> tuple[bytes | None, dict[str, Any]]:
    resolved = Path(resolved_font_path or font_path or "")
    default_resolved = DEFAULT_FONT.resolve() if DEFAULT_FONT.exists() else None
    if resolved and resolved.exists() and (default_resolved is None or resolved.resolve() != default_resolved):
        return resolved.read_bytes(), {
            "font_source": "custom_path",
            "font_file_name": resolved.name,
        }
    return None, {
        "font_source": "default" if DEFAULT_FONT.exists() else "none",
        "font_file_name": DEFAULT_FONT.name if DEFAULT_FONT.exists() else "",
    }


def _update_job_failed(job_id: str | None, exc: Exception) -> None:
    if not job_id:
        return
    structured = error_metadata(exc, step="translate_text")
    update_translation_job(
        job_id,
        status="failed",
        progress=0,
        message="翻译失败",
        error=str(exc),
        result_meta=json.dumps(
            {
                "error": str(exc),
                "error_code": structured.get("error_code"),
                "error_step": structured.get("error_step"),
                "retryable": structured.get("retryable"),
                "error_message": structured.get("error_message"),
                "traceback": traceback.format_exc()[:4000],
                "candidate_ids": [],
                "candidate_count": 0,
                "candidate_ids_reliable": False,
                "unrecorded_term_count": None,
                "review_item_count": None,
            },
            ensure_ascii=False,
        ),
    )


def _candidate_metadata_for_job(
    translation_job_id: str | None,
    review_summary: dict | None = None,
    reliable: bool = True,
) -> tuple[list[int], int, dict[str, Any]]:
    summary = review_summary or {}
    review_item_count = int(summary.get("n_review_items", 0) or 0)
    if not translation_job_id or not reliable:
        return [], 0, {
            "candidate_ids_reliable": False,
            "review_item_count": review_item_count,
        }
    rows = _core_list_term_candidates_for_job(translation_job_id)
    candidate_ids = [int(row["candidate_id"]) for row in rows]
    return candidate_ids, len(candidate_ids), {
        "candidate_ids_reliable": True,
        "review_item_count": review_item_count,
    }


def list_customers(user: dict | None = None, include_term_counts: bool = False) -> list[dict]:
    """Return all customers as plain dicts. `user` is reserved for future permission filtering."""
    _ensure_db_ready()
    if include_term_counts:
        rows = list_customers_with_term_counts({"role": "company_admin", "username": "service"})
        return [dict(row) for row in rows]
    return [dict(row) for row in _core_list_customers()]


def get_customer(customer_id: str, user: dict | None = None) -> dict | None:
    """Return one customer as a plain dict. `user` is reserved for future permission filtering."""
    _ensure_db_ready()
    customer = _core_get_customer(customer_id)
    return dict(customer) if customer else None


def detect_pdf_translation_scope(
    *,
    filename: str = "",
    file_bytes: bytes,
) -> dict[str, Any]:
    """Detect likely workmanship pages.

    `detected_pages` is 0-based for service/run_pdf_translation use.
    Each row in `scope_detection` also includes `page_number` as 1-based display value.
    """
    with fitz.open(stream=file_bytes, filetype="pdf") as doc:
        page_count = len(doc)
    scope_detection = detect_workmanship_pages(file_bytes)
    detected_pages = [
        int(row["page_number"]) - 1
        for row in scope_detection
        if row.get("is_workmanship") and row.get("page_number")
    ]
    return {
        "source_filename": filename,
        "page_count": page_count,
        "detected_pages": detected_pages,
        "display_page_numbers": [p + 1 for p in detected_pages],
        "scope_detection": scope_detection,
    }


def detect_excel_translation_scope(
    *,
    filename: str = "",
    file_bytes: bytes,
) -> dict[str, Any]:
    """Detect likely workmanship sheets. Sheet names are used directly, not indexed."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        all_sheets = list(wb.sheetnames)
    finally:
        wb.close()
    scope_detection = detect_workmanship_sheets(file_bytes)
    detected_sheets = [
        str(row["sheet_name"])
        for row in scope_detection
        if row.get("is_workmanship") and row.get("sheet_name")
    ]
    return {
        "source_filename": filename,
        "all_sheets": all_sheets,
        "detected_sheets": detected_sheets,
        "scope_detection": scope_detection,
    }


def _detect_file_type(filename: str, file_bytes: bytes, file_type: str | None = None) -> str:
    ext = (file_type or Path(filename or "").suffix.lstrip(".")).lower().strip()
    if ext in {"pdf", "xlsx", "xls"}:
        return ext
    if file_bytes[:4] == b"%PDF":
        return "pdf"
    if file_bytes[:8] == b"PK\x03\x04":
        return "xlsx"
    if file_bytes[:4] == b"\xd0\xcf\x11\xe0":
        return "xls"
    raise ValueError("无法识别文件类型，请提供 pdf / xlsx / xls 文件")


def _resolve_default_font_path(font_path: str | None) -> str:
    resolved = font_path or (str(DEFAULT_FONT) if DEFAULT_FONT.exists() else "")
    if not resolved:
        raise FileNotFoundError("未找到默认字体，请提供 font_path")
    return resolved


def _load_customer_glossary_bytes(customer_id: str) -> bytes:
    glossary_df = get_customer_glossary_df(customer_id)
    return glossary_df_to_xlsx_bytes(glossary_df)


def _convert_xls_to_xlsx_bytes(file_bytes: bytes, filename: str) -> tuple[bytes, dict[str, Any]]:
    sheet_rows, parse_error = _read_excel_sheet_rows(file_bytes, filename)
    if parse_error:
        raise ValueError(parse_error)

    wb = openpyxl.Workbook()
    first = True
    sheet_names: list[str] = []
    for sheet_name, rows in sheet_rows:
        sheet_names.append(sheet_name)
        if first:
            ws = wb.active
            ws.title = sheet_name[:31] or "Sheet1"
            first = False
        else:
            ws = wb.create_sheet(title=sheet_name[:31] or f"Sheet{len(sheet_names)}")
        for row in rows:
            ws.append([cell if cell is not None else "" for cell in row])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), {
        "excel_conversion": "xls_to_xlsx_values_only",
        "xls_sheet_names": sheet_names,
    }


def _build_pdf_report_bytes(
    source_filename: str,
    total_pages: int,
    selected_pages,
    scope_detection,
    scope_mode: str,
    glossary_report_bytes: bytes,
) -> tuple[bytes, dict[str, Any]]:
    scope_cfg = {
        "total_pages": total_pages,
        "detection": scope_detection or [],
    }
    scope_report = _pdf_scope_report_rows(
        source_filename,
        scope_cfg,
        selected_pages,
        scope_detection,
        scope_mode,
    )
    scope_report_bytes = build_scope_report_xlsx(scope_report) if scope_report else b""
    report_bytes = b""
    report_kind = ""
    if glossary_report_bytes and scope_report_bytes:
        report_zip = io.BytesIO()
        with zipfile.ZipFile(report_zip, "w") as zf:
            zf.writestr("unrecorded_terms.xlsx", glossary_report_bytes)
            zf.writestr("scope_report.xlsx", scope_report_bytes)
        report_bytes = report_zip.getvalue()
        report_kind = "zip"
    elif glossary_report_bytes:
        report_bytes = glossary_report_bytes
        report_kind = "unrecorded"
    elif scope_report_bytes:
        report_bytes = scope_report_bytes
        report_kind = "scope"

    meta = {
        "has_scope_report": bool(scope_report),
        "has_unrecorded_terms": bool(glossary_report_bytes),
        "report_kind": report_kind,
    }
    return report_bytes, meta


def _build_excel_report_bundle(
    report_rows: list[dict],
    img_report_rows: list[dict] | None = None,
) -> bytes:
    report_wb = openpyxl.Workbook()
    report_ws = report_wb.active
    report_ws.title = "翻译报告"
    headers = [
        "sheet_name", "cell_coordinate", "original_text", "translated_text",
        "status", "skip_reason", "is_merged_cell", "layout_warning",
        "scope_mode", "selected_sheets", "skipped_sheets",
        "detection_score", "detection_reason",
    ]
    report_ws.append(headers)
    for cell in report_ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for r in report_rows:
        report_ws.append([r.get(h, "") for h in headers])
    for col_letter, width in zip(
        "ABCDEFGHIJKLM",
        [12, 14, 35, 35, 12, 24, 14, 14, 18, 35, 35, 12, 70],
    ):
        report_ws.column_dimensions[col_letter].width = width

    if img_report_rows:
        img_ws = report_wb.create_sheet("图片译文")
        img_headers = ["drawing", "image", "status", "original_text", "translated_text", "skip_reason"]
        img_ws.append(img_headers)
        for cell in img_ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        for r in img_report_rows:
            img_ws.append([r.get(h, "") for h in img_headers])
        for col_letter, width in zip("ABCDEF", [25, 18, 10, 35, 35, 20]):
            img_ws.column_dimensions[col_letter].width = width

    buf = io.BytesIO()
    report_wb.save(buf)
    return buf.getvalue()


def translate_pdf_document(
    *,
    customer_id: str,
    filename: str,
    file_bytes: bytes,
    api_key: str,
    user: dict | None = None,
    font_path: str | None = None,
    selected_pages: list[int] | None = None,
    scope_mode: str = "all",
    scope_detection: list[dict] | None = None,
    source_file_name: str | None = None,
    created_by: str | None = None,
) -> dict[str, Any]:
    _ensure_db_ready()
    job_id: str | None = None
    source_name = _safe_filename(filename, source_file_name)
    try:
        if not created_by:
            created_by = user.get("username") if isinstance(user, dict) and user.get("username") else "service"
        job_config = {
            "service_entrypoint": "translation_service.translate_pdf_document",
            "customer_id": customer_id,
            "source_file_name": source_name,
            "selected_pages": selected_pages,
            "scope_mode": scope_mode,
            "scope_detection": scope_detection or [],
            "font": {
                "font_source": "custom_path" if font_path else ("default" if DEFAULT_FONT.exists() else "none"),
                "font_file_name": Path(font_path).name if font_path else (DEFAULT_FONT.name if DEFAULT_FONT.exists() else ""),
            },
        }
        job_id = create_translation_job(
            job_type="PDF",
            username=created_by,
            customer_id=customer_id,
            source_file_name=source_name,
            input_bytes=file_bytes,
            aux_bytes=None,
            config=job_config,
            execution_mode="sync",
            initial_status="running",
            initial_progress=0.01,
            initial_message="正在准备翻译",
        )
        if user is not None and not can_use_customer_glossary(user, customer_id):
            raise PermissionError("当前用户无权使用该客户术语库")
        glossary_bytes = _load_customer_glossary_bytes(customer_id)
        resolved_font_path = _resolve_default_font_path(font_path)
        with fitz.open(stream=file_bytes, filetype="pdf") as doc:
            input_page_count = len(doc)
        font_bytes, font_meta = _font_job_payload(font_path, resolved_font_path)
        job_config["font"] = font_meta
        update_translation_job(job_id, aux_bytes=font_bytes, config=json.dumps(job_config, ensure_ascii=False))

        def on_page(pn, total, n_blocks):
            update_translation_job(job_id, message=f"第 {pn + 1}/{total} 页，{n_blocks} 个文本块")

        def on_block(preview):
            update_translation_job(job_id, message=str(preview)[:180])

        def on_progress(frac):
            update_translation_job(job_id, progress=max(0.0, min(float(frac), 1.0)))

        pdf_out, xlsx_out, n_terms, review_summary = run_pdf_translation(
            pdf_bytes=file_bytes,
            glossary_bytes=glossary_bytes,
            font_path=resolved_font_path,
            api_key=api_key,
            on_page=on_page,
            on_block=on_block,
            on_progress=on_progress,
            customer_id=customer_id,
            source_file_name=source_name,
            created_by=created_by,
            selected_pages=selected_pages,
            scope_mode=scope_mode,
            scope_detection=scope_detection,
            translation_job_id=job_id,
        )
        with fitz.open(stream=pdf_out, filetype="pdf") as out_doc:
            output_page_count = len(out_doc)
        report_bytes, report_meta = _build_pdf_report_bytes(
            source_filename=source_name,
            total_pages=input_page_count,
            selected_pages=selected_pages,
            scope_detection=scope_detection,
            scope_mode=scope_mode,
            glossary_report_bytes=xlsx_out,
        )
        candidate_ids, candidate_count, candidate_meta = _candidate_metadata_for_job(job_id, review_summary)
        unrecorded_terms = list((review_summary or {}).get("unrecorded_terms") or [])
        meta = {
            "n_terms": n_terms,
            "unrecorded_term_count": n_terms,
            "unrecorded_terms": unrecorded_terms,
            "review_summary": review_summary,
            "customer_id": customer_id,
            "input_page_count": input_page_count,
            "output_page_count": output_page_count,
            "selected_pages": selected_pages,
            "scope_mode": scope_mode,
            "candidate_ids": candidate_ids,
            "candidate_count": candidate_count,
            **candidate_meta,
            **report_meta,
        }
        update_translation_job(
            job_id,
            status="complete",
            progress=1.0,
            message="翻译完成",
            error="",
            result_file=pdf_out,
            result_report=report_bytes,
            result_meta=json.dumps(meta, ensure_ascii=False),
        )
        return {
            "status": "completed",
            "translation_job_id": job_id,
            "customer_id": customer_id,
            "source_filename": source_name,
            "file_type": "pdf",
            "translated_file_bytes": pdf_out,
            "report_bytes": report_bytes,
            "candidate_ids": candidate_ids,
            "candidate_count": candidate_count,
            "unrecorded_term_count": n_terms,
            "unrecorded_terms": unrecorded_terms,
            "meta": meta,
            "error": None,
        }
    except Exception as exc:
        _update_job_failed(job_id, exc)
        return {
            "status": "failed",
            "translation_job_id": job_id,
            "customer_id": customer_id,
            "source_filename": source_name,
            "file_type": "pdf",
            "translated_file_bytes": None,
            "report_bytes": None,
            "candidate_ids": [],
            "candidate_count": 0,
            "unrecorded_term_count": 0,
            "unrecorded_terms": [],
            "meta": {"traceback": traceback.format_exc()},
            "error": str(exc),
        }


def translate_excel_document(
    *,
    customer_id: str,
    filename: str,
    file_bytes: bytes,
    api_key: str,
    user: dict | None = None,
    translate_images: bool = False,
    selected_sheets: list[str] | None = None,
    scope_mode: str = "all",
    scope_detection: list[dict] | None = None,
    source_file_name: str | None = None,
    created_by: str | None = None,
) -> dict[str, Any]:
    _ensure_db_ready()
    job_id: str | None = None
    source_name = _safe_filename(filename, source_file_name)
    try:
        if not created_by:
            created_by = user.get("username") if isinstance(user, dict) and user.get("username") else "service"
        job_config = {
            "service_entrypoint": "translation_service.translate_excel_document",
            "customer_id": customer_id,
            "source_file_name": source_name,
            "selected_sheets": selected_sheets,
            "scope_mode": scope_mode,
            "scope_detection": scope_detection or [],
            "translate_images": translate_images,
        }
        job_id = create_translation_job(
            job_type="Excel",
            username=created_by,
            customer_id=customer_id,
            source_file_name=source_name,
            input_bytes=file_bytes,
            aux_bytes=None,
            config=job_config,
            execution_mode="sync",
            initial_status="running",
            initial_progress=0.01,
            initial_message="正在准备翻译",
        )
        if user is not None and not can_use_customer_glossary(user, customer_id):
            raise PermissionError("当前用户无权使用该客户术语库")
        if Path(filename or "").suffix.lower() == ".xls":
            raise ValueError("当前仅支持 .xlsx；.xls 请先转换为 .xlsx 后再翻译")
        glossary_bytes = _load_customer_glossary_bytes(customer_id)
        glossary_dict = load_glossary(glossary_bytes)
        working_bytes = file_bytes

        def on_cell(preview):
            update_translation_job(job_id, message=str(preview)[:180])

        def on_progress(frac):
            update_translation_job(job_id, progress=max(0.0, min(float(frac), 1.0)))

        excel_out, n_cells, n_images, report_rows, review_summary = run_excel_translation(
            xlsx_bytes=working_bytes,
            glossary_bytes=glossary_bytes,
            api_key=api_key,
            on_cell=on_cell,
            on_progress=on_progress,
            translate_images=False,
            customer_id=customer_id,
            source_file_name=source_name,
            created_by=created_by,
            selected_sheets=selected_sheets,
            scope_mode=scope_mode,
            scope_detection=scope_detection,
            translation_job_id=job_id,
        )
        img_report_rows: list[dict] = []
        if translate_images:
            client = OpenAI(api_key=api_key)
            excel_out, img_report_rows = add_translated_textboxes_to_excel(
                excel_out,
                client,
                glossary_dict,
                on_image=lambda i, total, fname: update_translation_job(
                    job_id,
                    message=f"图片 {i}/{total}：{fname}",
                    progress=0.6 + (float(i) / max(float(total), 1.0)) * 0.4,
                ),
                selected_sheets=selected_sheets,
            )
        report_bytes = _build_excel_report_bundle(report_rows, img_report_rows)
        translated_image_count = sum(1 for r in img_report_rows if r.get("status") == "ok")
        candidate_ids, candidate_count, candidate_meta = _candidate_metadata_for_job(job_id, review_summary)
        unrecorded_terms = list((review_summary or {}).get("unrecorded_terms") or [])
        unrecorded_term_count = int((review_summary or {}).get("unrecorded_term_count", len(unrecorded_terms)) or 0)
        meta = {
            "customer_id": customer_id,
            "selected_sheets": selected_sheets,
            "scope_mode": scope_mode,
            "translate_images": translate_images,
            "translated_cell_count": n_cells,
            "translated_image_count": translated_image_count,
            "unrecorded_term_count": unrecorded_term_count,
            "unrecorded_terms": unrecorded_terms,
            "image_translation_unrecorded_tracking_supported": False,
            "candidate_ids": candidate_ids,
            "candidate_count": candidate_count,
            "review_summary": review_summary,
            **candidate_meta,
        }
        update_translation_job(
            job_id,
            status="complete",
            progress=1.0,
            message="翻译完成",
            error="",
            result_file=excel_out,
            result_report=report_bytes,
            result_meta=json.dumps(meta, ensure_ascii=False),
        )
        return {
            "status": "completed",
            "translation_job_id": job_id,
            "customer_id": customer_id,
            "source_filename": source_name,
            "file_type": "excel",
            "translated_file_bytes": excel_out,
            "report_bytes": report_bytes,
            "candidate_ids": candidate_ids,
            "candidate_count": candidate_count,
            "unrecorded_term_count": unrecorded_term_count,
            "unrecorded_terms": unrecorded_terms,
            "meta": meta,
            "error": None,
        }
    except Exception as exc:
        _update_job_failed(job_id, exc)
        return {
            "status": "failed",
            "translation_job_id": job_id,
            "customer_id": customer_id,
            "source_filename": source_name,
            "file_type": "excel",
            "translated_file_bytes": None,
            "report_bytes": None,
            "candidate_ids": [],
            "candidate_count": 0,
            "unrecorded_term_count": 0,
            "unrecorded_terms": [],
            "meta": {"traceback": traceback.format_exc()},
            "error": str(exc),
        }


def translate_customer_document(
    *,
    customer_id: str,
    filename: str,
    file_bytes: bytes,
    api_key: str,
    user: dict | None = None,
    file_type: str | None = None,
    font_path: str | None = None,
    translate_images: bool = False,
    selected_pages: list[int] | None = None,
    selected_sheets: list[str] | None = None,
    scope_mode: str = "all",
    scope_detection: list[dict] | None = None,
    source_file_name: str | None = None,
    created_by: str | None = None,
) -> dict[str, Any]:
    kind = _detect_file_type(filename, file_bytes, file_type)
    if kind == "pdf":
        return translate_pdf_document(
            customer_id=customer_id,
            filename=filename,
            file_bytes=file_bytes,
            api_key=api_key,
            user=user,
            font_path=font_path,
            selected_pages=selected_pages,
            scope_mode=scope_mode,
            scope_detection=scope_detection,
            source_file_name=source_file_name,
            created_by=created_by,
        )
    if kind == "xls":
        return translate_excel_document(
            customer_id=customer_id,
            filename=filename,
            file_bytes=file_bytes,
            api_key=api_key,
            user=user,
            translate_images=translate_images,
            selected_sheets=selected_sheets,
            scope_mode=scope_mode,
            scope_detection=scope_detection,
            source_file_name=source_file_name,
            created_by=created_by,
        )
    if kind == "xlsx":
        return translate_excel_document(
            customer_id=customer_id,
            filename=filename,
            file_bytes=file_bytes,
            api_key=api_key,
            user=user,
            translate_images=translate_images,
            selected_sheets=selected_sheets,
            scope_mode=scope_mode,
            scope_detection=scope_detection,
            source_file_name=source_file_name,
            created_by=created_by,
        )
    return {
        "status": "failed",
        "translation_job_id": None,
        "customer_id": customer_id,
        "source_filename": source_file_name or filename,
        "file_type": kind,
        "translated_file_bytes": None,
        "report_bytes": None,
        "candidate_ids": [],
        "candidate_count": 0,
        "unrecorded_term_count": 0,
        "unrecorded_terms": [],
        "meta": {},
        "error": f"不支持的文件类型：{kind}",
    }


def _normalize_job_type_filter(job_type: str | None) -> str | None:
    if not job_type:
        return None
    normalized = str(job_type).strip().lower()
    if normalized == "pdf":
        return "PDF"
    if normalized in {"excel", "xlsx", "xls"}:
        return "Excel"
    return str(job_type).strip()


def _normalize_status_filter(status: str | None) -> str | None:
    if not status:
        return None
    normalized = str(status).strip().lower()
    if normalized == "completed":
        return "complete"
    return str(status).strip()


def list_translation_jobs(
    *,
    customer_id: str | None = None,
    job_type: str | None = None,
    status: str | None = None,
    username: str | None = None,
    limit: int = 100,
) -> list[dict]:
    """List translation_jobs without large byte columns."""
    _ensure_db_ready()
    where: list[str] = []
    params: list[Any] = []
    if customer_id:
        where.append("customer_id = ?")
        params.append(customer_id)
    normalized_job_type = _normalize_job_type_filter(job_type)
    if normalized_job_type:
        where.append("job_type = ?")
        params.append(normalized_job_type)
    normalized_status = _normalize_status_filter(status)
    if normalized_status:
        where.append("status = ?")
        params.append(normalized_status)
    if username:
        where.append("username = ?")
        params.append(username)
    where_sql = "WHERE " + " AND ".join(where) if where else ""
    with get_db_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT job_id, job_type, status, username, customer_id, source_file_name,
                   progress, message, error, result_meta, config, created_at, updated_at,
                   CASE WHEN result_file IS NULL THEN 0 ELSE length(result_file) END AS result_file_size,
                   CASE WHEN result_report IS NULL THEN 0 ELSE length(result_report) END AS result_report_size
            FROM translation_jobs
            {where_sql}
            ORDER BY created_at DESC
            LIMIT ?
            """,
            [*params, int(limit)],
        ).fetchall()
    jobs = []
    for row in rows:
        item = dict(row)
        item["result_meta"] = _normalize_result_meta(item.get("result_meta"))
        item["config"] = _json_loads_dict(item.get("config"))
        item["has_result_file"] = bool(item.get("result_file_size"))
        item["has_result_report"] = bool(item.get("result_report_size"))
        jobs.append(item)
    return jobs


def get_translation_job(job_id: str, include_input_bytes: bool = False) -> dict | None:
    """Return one translation job. Result bytes are included; input bytes are optional."""
    _ensure_db_ready()
    columns = """
        job_id, job_type, status, username, customer_id, source_file_name,
        progress, message, error, result_file, result_report, result_meta,
        config, created_at, updated_at
    """
    if include_input_bytes:
        columns += ", input_bytes, aux_bytes"
    with get_db_connection() as conn:
        row = conn.execute(
            f"SELECT {columns} FROM translation_jobs WHERE job_id = ?",
            (job_id,),
        ).fetchone()
    if row is None:
        return None
    job = dict(row)
    job["result_meta"] = _normalize_result_meta(job.get("result_meta"))
    job["config"] = _json_loads_dict(job.get("config"))
    return job


def delete_translation_job(job_id: str) -> bool:
    _ensure_db_ready()
    with get_db_connection() as conn:
        conn.execute(
            "DELETE FROM translation_candidate_occurrences WHERE translation_job_id = ?",
            (job_id,),
        )
        cur = conn.execute("DELETE FROM translation_jobs WHERE job_id = ?", (job_id,))
        return bool(getattr(cur, "rowcount", 0))


def list_term_candidates_for_job(translation_job_id: str) -> dict:
    _ensure_db_ready()
    candidates = [dict(row) for row in _core_list_term_candidates_for_job(translation_job_id)]
    job = get_translation_job(translation_job_id)
    meta = (job or {}).get("result_meta") or {}
    reliable = bool(meta.get("candidate_ids_reliable")) if job else False
    if candidates and not reliable:
        reliable = True
    return {
        "translation_job_id": translation_job_id,
        "candidate_ids_reliable": reliable,
        "candidate_count": len(candidates) if reliable else None,
        "candidates": candidates,
    }


def list_candidate_occurrences(candidate_id: int) -> list[dict]:
    _ensure_db_ready()
    return [dict(row) for row in _core_list_candidate_occurrences(int(candidate_id))]


def list_term_candidates(
    *,
    customer_id: str | None = None,
    status: str | None = None,
    source_file_name: str | None = None,
    limit: int = 500,
) -> list[dict]:
    _ensure_db_ready()
    where: list[str] = []
    params: list[Any] = []
    if customer_id:
        where.append("tc.customer_id = ?")
        params.append(customer_id)
    if status and status != "全部":
        where.append("tc.status = ?")
        params.append(status)
    if source_file_name:
        where.append("tc.source_file_name = ?")
        params.append(source_file_name)
    where_sql = "WHERE " + " AND ".join(where) if where else ""
    with get_db_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT tc.*, c.customer_code, c.customer_name
            FROM term_candidates tc
            LEFT JOIN customers c ON c.customer_id = tc.customer_id
            {where_sql}
            ORDER BY tc.created_at DESC, tc.candidate_id DESC
            LIMIT ?
            """,
            [*params, int(limit)],
        ).fetchall()
    return [dict(row) for row in rows]


def get_term_candidate(candidate_id: int) -> dict | None:
    _ensure_db_ready()
    with get_db_connection() as conn:
        row = conn.execute(
            """
            SELECT tc.*, c.customer_code, c.customer_name
            FROM term_candidates tc
            LEFT JOIN customers c ON c.customer_id = tc.customer_id
            WHERE tc.candidate_id = ?
            """,
            (int(candidate_id),),
        ).fetchone()
    return dict(row) if row else None


def update_term_candidate(
    candidate_id: int,
    *,
    final_translation: str | None = None,
    username: str | None = None,
) -> dict:
    _ensure_db_ready()
    if final_translation is None:
        return {
            "status": "failed",
            "candidate_id": int(candidate_id),
            "error": "没有提供可更新字段",
            "candidate": get_term_candidate(int(candidate_id)),
        }
    with get_db_connection() as conn:
        row = conn.execute(
            "SELECT * FROM term_candidates WHERE candidate_id = ?",
            (int(candidate_id),),
        ).fetchone()
        if row is None:
            return {
                "status": "failed",
                "candidate_id": int(candidate_id),
                "error": "candidate 不存在",
                "candidate": None,
            }
        if row["status"] not in {"draft", "selected"}:
            return {
                "status": "failed",
                "candidate_id": int(candidate_id),
                "error": f"当前状态 {row['status']} 不允许修改 final_translation",
                "candidate": dict(row),
            }
        conn.execute(
            """
            UPDATE term_candidates
            SET final_translation = ?
            WHERE candidate_id = ?
            """,
            (str(final_translation).strip(), int(candidate_id)),
        )
    return {
        "status": "completed",
        "candidate_id": int(candidate_id),
        "candidate": get_term_candidate(int(candidate_id)),
        "error": None,
    }


def ignore_term_candidates(
    candidate_ids: list,
    *,
    username: str | None = None,
) -> dict:
    _ensure_db_ready()
    ignored: list[int] = []
    failed: list[dict] = []
    for raw_id in candidate_ids:
        try:
            candidate_id = int(raw_id)
        except (TypeError, ValueError):
            failed.append({"candidate_id": raw_id, "error": "candidate_id 无效"})
            continue
        with get_db_connection() as conn:
            row = conn.execute(
                "SELECT candidate_id FROM term_candidates WHERE candidate_id = ?",
                (candidate_id,),
            ).fetchone()
            if row is None:
                failed.append({"candidate_id": candidate_id, "error": "candidate 不存在"})
                continue
            conn.execute(
                "UPDATE term_candidates SET status = 'ignored' WHERE candidate_id = ?",
                (candidate_id,),
            )
            ignored.append(candidate_id)
    status = "completed" if ignored and not failed else "partial" if ignored else "failed"
    return {
        "status": status,
        "ignored_candidate_ids": ignored,
        "failed": failed,
    }


def _submit_one_term_candidate_for_approval(
    item: dict,
    user: dict,
) -> dict:
    candidate_id = int(item.get("candidate_id"))
    final_translation = str(item.get("final_translation", "") or "").strip()
    note = str(item.get("comment", item.get("note", "")) or "").strip()
    if not final_translation:
        raise ValueError("缺少 final_translation")

    with get_db_connection() as conn:
        candidate = conn.execute(
            """
            SELECT *
            FROM term_candidates
            WHERE candidate_id = ? AND status IN ('draft', 'selected')
            """,
            (candidate_id,),
        ).fetchone()
        if candidate is None:
            raise ValueError("candidate 不存在，或当前状态不允许提交")

        customer_id = candidate["customer_id"]
        active_rows = conn.execute(
            "SELECT english_term FROM glossary_terms WHERE customer_id = ? AND status = 'active'",
            (customer_id,),
        ).fetchall()
        active_keys = {
            normalize_term_key(r["english_term"]) or normalize_term(r["english_term"])
            for r in active_rows
        }
        pending_keys = _pending_term_keys(customer_id)
        key = normalize_term_key(candidate["original_term"]) or normalize_term(candidate["original_term"])
        if key in active_keys:
            conn.execute(
                "UPDATE term_candidates SET status = 'ignored' WHERE candidate_id = ?",
                (candidate_id,),
            )
            raise ValueError(f"{candidate['original_term']} 已在正式术语库中，已忽略")
        if key in pending_keys:
            raise ValueError(f"{candidate['original_term']} 已有待审批申请")
        candidate = dict(candidate)

    source_note = (
        f"candidate_id={candidate_id}；来源文件={candidate['source_file_name']}；"
        f"来源位置={candidate['page_or_sheet']} {candidate['cell_coordinate'] or ''}；"
        f"normalized_key={candidate['normalized_term']}；"
        f"variants={candidate['variants'] or candidate['original_term']}；"
        f"出现次数={candidate['frequency']}；置信度={candidate['confidence']}；"
        f"上下文={candidate['context_sentence']}"
    )
    full_note = f"{note}；{source_note}" if note else source_note
    request_id = create_glossary_change_request(
        user,
        candidate["customer_id"],
        "add",
        english_term_new=candidate["original_term"],
        chinese_translation_new=final_translation,
        note=full_note,
        candidate_id=candidate_id,
    )
    with get_db_connection() as conn:
        conn.execute(
            """
            UPDATE term_candidates
            SET final_translation = ?, status = 'submitted', submitted_at = ?
            WHERE candidate_id = ?
            """,
            (final_translation, _now_iso(), candidate_id),
        )
    return {
        "candidate_id": candidate_id,
        "request_id": int(request_id),
    }


def submit_term_candidates_for_approval(
    rows: list[dict],
    *,
    username: str | None = None,
) -> dict:
    _ensure_db_ready()
    user = _service_user(username)
    submitted_candidate_ids: list[int] = []
    request_ids: list[int] = []
    failed: list[dict] = []
    for row in rows:
        raw_id = row.get("candidate_id")
        try:
            result = _submit_one_term_candidate_for_approval(row, user)
            submitted_candidate_ids.append(result["candidate_id"])
            request_ids.append(result["request_id"])
        except Exception as exc:
            failed.append({"candidate_id": raw_id, "error": str(exc)})
    status = "completed" if submitted_candidate_ids and not failed else "partial" if submitted_candidate_ids else "failed"
    return {
        "status": status,
        "submitted_candidate_ids": submitted_candidate_ids,
        "request_ids": request_ids,
        "failed": failed,
    }


def _glossary_row_to_dict(row: Any) -> dict[str, Any]:
    data = dict(row)
    return {
        "glossary_id": data.get("glossary_id"),
        "customer_id": data.get("customer_id"),
        "english_term": data.get("english_term") or "",
        "chinese_translation": data.get("chinese_translation") or "",
        "normalized_key": data.get("normalized_key") or "",
        "note": data.get("note") or "",
        "status": data.get("status") or "",
        "created_by": data.get("created_by") or "",
        "updated_by": data.get("updated_by") or "",
        "updated_at": data.get("updated_at") or "",
    }


def list_customer_glossary(
    customer_id: str,
    *,
    status: str | None = "active",
    limit: int = 1000,
    offset: int = 0,
) -> list[dict]:
    """Return glossary terms for one customer as plain dicts."""
    _ensure_db_ready()
    where = ["customer_id = ?"]
    params: list[Any] = [customer_id]
    if status and str(status).lower() not in {"all", "全部"}:
        where.append("status = ?")
        params.append(str(status))
    with get_db_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT glossary_id, customer_id, english_term, chinese_translation,
                   normalized_key, note, status, created_by, updated_by, updated_at
            FROM glossary_terms
            WHERE {' AND '.join(where)}
            ORDER BY lower(english_term), glossary_id
            LIMIT ? OFFSET ?
            """,
            [*params, int(limit), int(offset)],
        ).fetchall()
    return [_glossary_row_to_dict(row) for row in rows]


def search_customer_glossary(
    customer_id: str,
    keyword: str,
    *,
    status: str | None = "active",
    limit: int = 100,
) -> list[dict]:
    """Search one customer's glossary by English term or Chinese translation."""
    _ensure_db_ready()
    pattern = f"%{str(keyword or '').strip()}%"
    where = ["customer_id = ?", "(english_term LIKE ? OR chinese_translation LIKE ?)"]
    params: list[Any] = [customer_id, pattern, pattern]
    if status and str(status).lower() not in {"all", "全部"}:
        where.append("status = ?")
        params.append(str(status))
    with get_db_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT glossary_id, customer_id, english_term, chinese_translation,
                   normalized_key, note, status, created_by, updated_by, updated_at
            FROM glossary_terms
            WHERE {' AND '.join(where)}
            ORDER BY lower(english_term), glossary_id
            LIMIT ?
            """,
            [*params, int(limit)],
        ).fetchall()
    return [_glossary_row_to_dict(row) for row in rows]


def get_glossary_term(glossary_id: int) -> dict | None:
    _ensure_db_ready()
    with get_db_connection() as conn:
        row = conn.execute(
            """
            SELECT glossary_id, customer_id, english_term, chinese_translation,
                   normalized_key, note, status, created_by, updated_by, updated_at
            FROM glossary_terms
            WHERE glossary_id = ?
            """,
            (int(glossary_id),),
        ).fetchone()
    return _glossary_row_to_dict(row) if row else None


def export_customer_glossary(customer_id: str) -> bytes:
    """Export active glossary terms in the existing PDF Project workbook format."""
    _ensure_db_ready()
    return glossary_df_to_xlsx_bytes(get_customer_glossary_df(customer_id))


def _active_glossary_maps(customer_id: str) -> tuple[dict[str, dict], dict[str, list[dict]]]:
    with get_db_connection() as conn:
        rows = conn.execute(
            """
            SELECT glossary_id, customer_id, english_term, chinese_translation,
                   normalized_key, note, status, created_by, updated_by, updated_at
            FROM glossary_terms
            WHERE customer_id = ? AND status = 'active'
            """,
            (customer_id,),
        ).fetchall()
    by_english = {str(r["english_term"]).strip().lower(): dict(r) for r in rows}
    by_key: dict[str, list[dict]] = {}
    for row in rows:
        key = row["normalized_key"] or normalize_term_key(row["english_term"]) or normalize_term(row["english_term"])
        if key:
            by_key.setdefault(key, []).append(dict(row))
    return by_english, by_key


def _pending_request_rows(customer_id: str) -> list[dict]:
    with get_db_connection() as conn:
        rows = conn.execute(
            """
            SELECT request_id, customer_id, action_type, english_term_old,
                   chinese_translation_old, english_term_new, chinese_translation_new,
                   note, submitted_by, submitted_at, status
            FROM glossary_change_requests
            WHERE customer_id = ? AND status = 'pending'
            ORDER BY submitted_at DESC, request_id DESC
            """,
            (customer_id,),
        ).fetchall()
    return [dict(row) for row in rows]


def _pending_request_matches(customer_id: str, action_type: str, english_term: str) -> list[dict]:
    key = normalize_term_key(english_term) or normalize_term(english_term)
    matches: list[dict] = []
    for row in _pending_request_rows(customer_id):
        if row.get("action_type") != action_type:
            continue
        term = row.get("english_term_new") if action_type in {"add", "update"} else row.get("english_term_old")
        row_key = normalize_term_key(term or "") or normalize_term(term or "")
        if key and row_key == key:
            matches.append(row)
    return matches


def _pending_add_update_matches(customer_id: str, english_term: str) -> list[dict]:
    key = normalize_term_key(english_term) or normalize_term(english_term)
    matches: list[dict] = []
    for row in _pending_request_rows(customer_id):
        if row.get("action_type") not in {"add", "update"}:
            continue
        row_key = normalize_term_key(row.get("english_term_new") or "") or normalize_term(row.get("english_term_new") or "")
        if key and row_key == key:
            matches.append(row)
    return matches


def _active_key_conflict(
    by_key: dict[str, list[dict]],
    normalized_key: str,
    chinese_translation: str,
    *,
    ignore_glossary_id: int | None = None,
) -> bool:
    rows = by_key.get(normalized_key, []) if normalized_key else []
    if ignore_glossary_id is not None:
        rows = [r for r in rows if int(r["glossary_id"]) != int(ignore_glossary_id)]
    return glossary_key_has_translation_conflict(rows, chinese_translation)


def request_add_glossary_term(
    *,
    customer_id: str,
    english_term: str,
    chinese_translation: str,
    note: str | None = None,
    submitted_by: str | None = None,
) -> dict:
    """Create a pending add request; never inserts directly into active glossary_terms."""
    _ensure_db_ready()
    en = str(english_term or "").strip()
    zh = str(chinese_translation or "").strip()
    if not en or not zh:
        return {"status": "failed", "request_id": None, "error": "英文术语和中文翻译不能为空"}
    normalized_key = normalize_term_key(en) or normalize_term(en)
    by_english, by_key = _active_glossary_maps(customer_id)
    if en.lower() in by_english:
        return {"status": "failed", "request_id": None, "error": "该客户已存在相同英文术语"}
    if _active_key_conflict(by_key, normalized_key, zh):
        return {"status": "failed", "request_id": None, "error": "同一 normalized_key 已存在不同中文翻译，请人工处理"}
    pending = _pending_add_update_matches(customer_id, en)
    if pending:
        return {
            "status": "failed",
            "request_id": None,
            "error": f"已有相同术语待审批申请 #{pending[0]['request_id']}",
        }
    request_id = create_glossary_change_request(
        _service_user(submitted_by),
        customer_id,
        "add",
        english_term_new=en,
        chinese_translation_new=zh,
        note=note or "",
    )
    return {
        "status": "completed",
        "request_id": int(request_id),
        "action_type": "add",
        "customer_id": customer_id,
        "normalized_key": normalized_key,
        "error": None,
    }


def request_update_glossary_term(
    *,
    glossary_id: int,
    english_term: str | None = None,
    chinese_translation: str | None = None,
    note: str | None = None,
    submitted_by: str | None = None,
) -> dict:
    """Create a pending update request; active glossary_terms stay unchanged."""
    _ensure_db_ready()
    old = get_glossary_term(int(glossary_id))
    if not old or old.get("status") != "active":
        return {"status": "failed", "request_id": None, "error": "active glossary term 不存在"}
    en_new = str(english_term if english_term is not None else old["english_term"]).strip()
    zh_new = str(chinese_translation if chinese_translation is not None else old["chinese_translation"]).strip()
    note_new = str(note if note is not None else old.get("note", "")).strip()
    if not en_new or not zh_new:
        return {"status": "failed", "request_id": None, "error": "英文术语和中文翻译不能为空"}
    if (
        en_new == old["english_term"]
        and zh_new == old["chinese_translation"]
        and note_new == (old.get("note") or "")
    ):
        return {"status": "failed", "request_id": None, "error": "术语内容无变化"}
    normalized_key = normalize_term_key(en_new) or normalize_term(en_new)
    _, by_key = _active_glossary_maps(old["customer_id"])
    if _active_key_conflict(by_key, normalized_key, zh_new, ignore_glossary_id=int(glossary_id)):
        return {"status": "failed", "request_id": None, "error": "同一 normalized_key 已存在不同中文翻译，请人工处理"}
    pending = _pending_add_update_matches(old["customer_id"], en_new)
    if pending:
        return {
            "status": "failed",
            "request_id": None,
            "error": f"已有相同术语待审批修改申请 #{pending[0]['request_id']}",
        }
    request_id = create_glossary_change_request(
        _service_user(submitted_by),
        old["customer_id"],
        "update",
        english_term_old=old["english_term"],
        chinese_translation_old=old["chinese_translation"],
        english_term_new=en_new,
        chinese_translation_new=zh_new,
        note=note_new,
    )
    return {
        "status": "completed",
        "request_id": int(request_id),
        "action_type": "update",
        "customer_id": old["customer_id"],
        "glossary_id": int(glossary_id),
        "normalized_key": normalized_key,
        "error": None,
    }


def request_delete_glossary_term(
    *,
    glossary_id: int,
    submitted_by: str | None = None,
    comment: str | None = None,
) -> dict:
    """Create a pending delete request; active glossary_terms stay unchanged."""
    _ensure_db_ready()
    old = get_glossary_term(int(glossary_id))
    if not old or old.get("status") != "active":
        return {"status": "failed", "request_id": None, "error": "active glossary term 不存在"}
    pending = _pending_request_matches(old["customer_id"], "delete", old["english_term"])
    if pending:
        return {
            "status": "failed",
            "request_id": None,
            "error": f"已有相同术语待审批删除申请 #{pending[0]['request_id']}",
        }
    request_id = create_glossary_change_request(
        _service_user(submitted_by),
        old["customer_id"],
        "delete",
        english_term_old=old["english_term"],
        chinese_translation_old=old["chinese_translation"],
        note=comment or "",
    )
    return {
        "status": "completed",
        "request_id": int(request_id),
        "action_type": "delete",
        "customer_id": old["customer_id"],
        "glossary_id": int(glossary_id),
        "error": None,
    }


def preview_customer_glossary_import(
    *,
    customer_id: str,
    file_bytes: bytes,
    filename: str = "",
) -> dict:
    """Parse and classify a glossary workbook without writing to the database."""
    _ensure_db_ready()
    df, missing, parse_stats, parse_report = parse_customer_glossary_excel(file_bytes, filename)
    preview_rows: list[dict] = []
    stats = {
        "add": 0,
        "update": 0,
        "no_change": 0,
        "error": int(parse_stats.get("error_rows", 0) or 0),
        "pending_duplicate": 0,
    }
    if missing:
        return {
            "status": "failed",
            "customer_id": customer_id,
            "filename": filename,
            "parse_errors": missing,
            "parse_stats": parse_stats,
            "summary": stats,
            "rows": preview_rows,
            "report_rows": parse_report,
            "report_bytes": build_import_report_xlsx(parse_report) if parse_report else b"",
            "error": "；".join(missing),
        }

    by_english, by_key = _active_glossary_maps(customer_id)
    pending_keys = _pending_term_keys(customer_id)
    for idx, row in df.iterrows():
        en = str(row.get(_GLOSSARY_EN_COL, "")).strip()
        zh = str(row.get(_GLOSSARY_ZH_COL, "")).strip()
        note = str(row.get(_GLOSSARY_NOTE_COL, "") or "").strip()
        source_sheet = str(row.get(_GLOSSARY_CAT_COL, "") or "").strip()
        normalized_key = normalize_term_key(en) or normalize_term(en)
        old = by_english.get(en.lower())
        action = "add"
        reason = "新增术语"
        glossary_id = old.get("glossary_id") if old else None
        if not en or not zh:
            action = "error"
            reason = "英文术语或中文翻译为空"
        elif normalized_key in pending_keys:
            action = "pending_duplicate"
            reason = "已有相同 normalized_key 的 pending 申请"
        elif old:
            if zh == old["chinese_translation"] and note == (old.get("note") or ""):
                action = "no_change"
                reason = "与 active 术语一致"
            else:
                action = "update"
                reason = "更新已有术语"
        elif _active_key_conflict(by_key, normalized_key, zh):
            action = "error"
            reason = "同一 normalized_key 已存在不同中文翻译，请人工处理"

        stats[action] = stats.get(action, 0) + 1
        preview_rows.append({
            "row_number": int(idx) + 2,
            "source_sheet": source_sheet,
            "action": action,
            "reason": reason,
            "glossary_id": glossary_id,
            "customer_id": customer_id,
            "english_term": en,
            "chinese_translation": zh,
            "normalized_key": normalized_key,
            "note": note,
            "old_english_term": old.get("english_term") if old else "",
            "old_chinese_translation": old.get("chinese_translation") if old else "",
            "old_note": old.get("note") if old else "",
        })

    report_rows = list(parse_report)
    report_rows.extend({
        "sheet_name": row["source_sheet"],
        "row_number": row["row_number"],
        "english_term": row["english_term"],
        "chinese_translation": row["chinese_translation"],
        "status": row["action"],
        "reason": row["reason"],
    } for row in preview_rows)
    return {
        "status": "completed",
        "customer_id": customer_id,
        "filename": filename,
        "parse_errors": [],
        "parse_stats": parse_stats,
        "summary": stats,
        "rows": preview_rows,
        "report_rows": report_rows,
        "report_bytes": build_import_report_xlsx(report_rows),
        "error": None,
    }


def request_import_customer_glossary(
    *,
    customer_id: str,
    file_bytes: bytes,
    filename: str = "",
    submitted_by: str | None = None,
) -> dict:
    """Create pending add/update requests from a glossary workbook; no active rows are changed."""
    _ensure_db_ready()
    preview = preview_customer_glossary_import(
        customer_id=customer_id,
        file_bytes=file_bytes,
        filename=filename,
    )
    if preview.get("status") != "completed":
        return {
            "status": "failed",
            "customer_id": customer_id,
            "filename": filename,
            "request_ids": [],
            "submitted_rows": [],
            "failed": [{"row_number": None, "error": preview.get("error") or "导入预览失败"}],
            "preview": preview,
        }

    request_ids: list[int] = []
    submitted_rows: list[dict] = []
    failed: list[dict] = []
    for row in preview["rows"]:
        try:
            if row["action"] == "add":
                result = request_add_glossary_term(
                    customer_id=customer_id,
                    english_term=row["english_term"],
                    chinese_translation=row["chinese_translation"],
                    note=row["note"],
                    submitted_by=submitted_by,
                )
            elif row["action"] == "update":
                result = request_update_glossary_term(
                    glossary_id=int(row["glossary_id"]),
                    english_term=row["english_term"],
                    chinese_translation=row["chinese_translation"],
                    note=row["note"],
                    submitted_by=submitted_by,
                )
            else:
                continue
            if result.get("status") == "completed":
                request_ids.append(int(result["request_id"]))
                submitted_rows.append({**row, "request_id": int(result["request_id"])})
            else:
                failed.append({
                    "row_number": row["row_number"],
                    "english_term": row["english_term"],
                    "action": row["action"],
                    "error": result.get("error") or "创建申请失败",
                })
        except Exception as exc:
            failed.append({
                "row_number": row.get("row_number"),
                "english_term": row.get("english_term"),
                "action": row.get("action"),
                "error": str(exc),
            })
    status = "completed" if request_ids and not failed else "partial" if request_ids else "failed"
    return {
        "status": status,
        "customer_id": customer_id,
        "filename": filename,
        "request_ids": request_ids,
        "submitted_rows": submitted_rows,
        "failed": failed,
        "summary": {
            "requested": len(request_ids),
            "failed": len(failed),
            "skipped_no_change": int(preview["summary"].get("no_change", 0) or 0),
            "skipped_errors": int(preview["summary"].get("error", 0) or 0),
            "skipped_pending_duplicate": int(preview["summary"].get("pending_duplicate", 0) or 0),
        },
        "preview": preview,
    }


def _change_request_row_to_dict(row: Any) -> dict[str, Any]:
    data = dict(row)
    data["submit_comment"] = data.get("note") or ""
    return data


def list_glossary_change_requests(
    *,
    customer_id: str | None = None,
    status: str | None = None,
    action_type: str | None = None,
    submitted_by: str | None = None,
    limit: int | None = None,
) -> list[dict]:
    """List glossary change requests as plain dicts."""
    _ensure_db_ready()
    where: list[str] = []
    params: list[Any] = []
    if customer_id:
        where.append("r.customer_id = ?")
        params.append(customer_id)
    if status:
        where.append("r.status = ?")
        params.append(status)
    if action_type:
        where.append("r.action_type = ?")
        params.append(action_type)
    if submitted_by:
        where.append("r.submitted_by = ?")
        params.append(submitted_by)
    where_sql = "WHERE " + " AND ".join(where) if where else ""
    limit_sql = " LIMIT ?" if limit is not None else ""
    if limit is not None:
        params.append(int(limit))
    with get_db_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT r.request_id, r.customer_id, r.action_type, r.candidate_id,
                   r.english_term_old, r.chinese_translation_old,
                   r.english_term_new, r.chinese_translation_new,
                   r.note, r.submitted_by, r.submitted_at, r.status,
                   r.reviewed_by, r.reviewed_at, r.review_comment,
                   c.customer_code, c.customer_name
            FROM glossary_change_requests r
            LEFT JOIN customers c ON c.customer_id = r.customer_id
            {where_sql}
            ORDER BY r.submitted_at DESC, r.request_id DESC
            {limit_sql}
            """,
            params,
        ).fetchall()
    return [_change_request_row_to_dict(row) for row in rows]


def get_glossary_change_request(request_id: int) -> dict | None:
    """Return one change request with a lightweight candidate summary when applicable."""
    _ensure_db_ready()
    with get_db_connection() as conn:
        row = conn.execute(
            """
            SELECT r.request_id, r.customer_id, r.action_type, r.candidate_id,
                   r.english_term_old, r.chinese_translation_old,
                   r.english_term_new, r.chinese_translation_new,
                   r.note, r.submitted_by, r.submitted_at, r.status,
                   r.reviewed_by, r.reviewed_at, r.review_comment,
                   c.customer_code, c.customer_name
            FROM glossary_change_requests r
            LEFT JOIN customers c ON c.customer_id = r.customer_id
            WHERE r.request_id = ?
            """,
            (int(request_id),),
        ).fetchone()
        if row is None:
            return None
        request = _change_request_row_to_dict(row)
        candidate_id = request.get("candidate_id")
        if candidate_id:
            candidate = conn.execute(
                """
                SELECT candidate_id, customer_id, original_term, normalized_term,
                       ai_suggested_translation, final_translation, status,
                       source_file_name, source_type, page_or_sheet, cell_coordinate,
                       context_sentence, frequency, confidence
                FROM term_candidates
                WHERE candidate_id = ?
                """,
                (candidate_id,),
            ).fetchone()
            request["candidate"] = dict(candidate) if candidate else None
        else:
            request["candidate"] = None
    return request


def _affected_glossary_id_for_request(request: dict | None) -> int | None:
    if not request:
        return None
    customer_id = request.get("customer_id")
    action = request.get("action_type")
    if action in {"add", "update"}:
        term = request.get("english_term_new")
        status_filter = "active"
    elif action == "delete":
        term = request.get("english_term_old")
        status_filter = "inactive"
    else:
        return None
    if not customer_id or not term:
        return None
    with get_db_connection() as conn:
        row = conn.execute(
            """
            SELECT glossary_id
            FROM glossary_terms
            WHERE customer_id = ? AND lower(english_term) = lower(?) AND status = ?
            ORDER BY glossary_id DESC
            LIMIT 1
            """,
            (customer_id, term, status_filter),
        ).fetchone()
    return int(row["glossary_id"]) if row else None


def approve_glossary_change_request(
    request_id: int,
    *,
    reviewer_username: str,
    review_comment: str | None = None,
) -> dict:
    """Approve a pending glossary change after server-side company_admin verification."""
    _ensure_db_ready()
    try:
        reviewer = _require_company_admin(reviewer_username)
        before = get_glossary_change_request(int(request_id))
        if before is None:
            return {
                "status": "failed",
                "request_id": int(request_id),
                "request_status": None,
                "error": "申请不存在",
            }
        if before.get("status") != "pending":
            return {
                "status": "failed",
                "request_id": int(request_id),
                "request_status": before.get("status"),
                "customer_id": before.get("customer_id"),
                "action_type": before.get("action_type"),
                "reviewed_by": before.get("reviewed_by"),
                "error": "request is no longer pending",
            }
        _core_approve_glossary_change_request(int(request_id), reviewer, review_comment or "")
        after = get_glossary_change_request(int(request_id))
        return {
            "status": "completed",
            "request_id": int(request_id),
            "request_status": (after or {}).get("status"),
            "customer_id": (after or before).get("customer_id"),
            "action_type": (after or before).get("action_type"),
            "reviewed_by": (after or {}).get("reviewed_by"),
            "reviewed_at": (after or {}).get("reviewed_at"),
            "affected_glossary_id": _affected_glossary_id_for_request(after),
            "error": None,
        }
    except Exception as exc:
        current = get_glossary_change_request(int(request_id))
        return {
            "status": "failed",
            "request_id": int(request_id),
            "request_status": (current or {}).get("status"),
            "customer_id": (current or {}).get("customer_id"),
            "action_type": (current or {}).get("action_type"),
            "reviewed_by": (current or {}).get("reviewed_by"),
            "error": str(exc),
        }


def reject_glossary_change_request(
    request_id: int,
    *,
    reviewer_username: str,
    review_comment: str | None = None,
) -> dict:
    """Reject a pending glossary change after server-side company_admin verification."""
    _ensure_db_ready()
    try:
        reviewer = _require_company_admin(reviewer_username)
        before = get_glossary_change_request(int(request_id))
        if before is None:
            return {
                "status": "failed",
                "request_id": int(request_id),
                "request_status": None,
                "error": "申请不存在",
            }
        if before.get("status") != "pending":
            return {
                "status": "failed",
                "request_id": int(request_id),
                "request_status": before.get("status"),
                "customer_id": before.get("customer_id"),
                "action_type": before.get("action_type"),
                "reviewed_by": before.get("reviewed_by"),
                "error": "request is no longer pending",
            }
        _core_reject_glossary_change_request(int(request_id), reviewer, review_comment or "")
        after = get_glossary_change_request(int(request_id))
        return {
            "status": "completed",
            "request_id": int(request_id),
            "request_status": (after or {}).get("status"),
            "customer_id": (after or before).get("customer_id"),
            "action_type": (after or before).get("action_type"),
            "reviewed_by": (after or {}).get("reviewed_by"),
            "reviewed_at": (after or {}).get("reviewed_at"),
            "affected_glossary_id": None,
            "error": None,
        }
    except Exception as exc:
        current = get_glossary_change_request(int(request_id))
        return {
            "status": "failed",
            "request_id": int(request_id),
            "request_status": (current or {}).get("status"),
            "customer_id": (current or {}).get("customer_id"),
            "action_type": (current or {}).get("action_type"),
            "reviewed_by": (current or {}).get("reviewed_by"),
            "error": str(exc),
        }


def translate_customer_document_unified(
    *,
    rfq_access_token: str,
    pdf_customer_id: str,
    filename: str,
    file_bytes: bytes,
    api_key: str,
    file_type: str | None = None,
    font_path: str | None = None,
    translate_images: bool = False,
    selected_pages: list[int] | None = None,
    selected_sheets: list[str] | None = None,
    scope_mode: str = "all",
    scope_detection: list[dict] | None = None,
    source_file_name: str | None = None,
) -> dict[str, Any]:
    principal, rfq_customer = _authorize_pdf_customer_unified(
        rfq_access_token,
        pdf_customer_id,
        require_quote_use=True,
    )
    result = translate_customer_document(
        customer_id=pdf_customer_id,
        filename=filename,
        file_bytes=file_bytes,
        api_key=api_key,
        user=None,
        file_type=file_type,
        font_path=font_path,
        translate_images=translate_images,
        selected_pages=selected_pages,
        selected_sheets=selected_sheets,
        scope_mode=scope_mode,
        scope_detection=scope_detection,
        source_file_name=source_file_name,
        created_by=principal["username"],
    )
    result.setdefault("meta", {})
    result["meta"]["unified_principal"] = {
        "user_id": principal["user_id"],
        "username": principal["username"],
        "role": principal["role"],
    }
    result["meta"]["rfq_customer"] = rfq_customer
    return result


def translate_pdf_document_unified(**kwargs) -> dict[str, Any]:
    return translate_customer_document_unified(file_type="pdf", **kwargs)


def translate_excel_document_unified(**kwargs) -> dict[str, Any]:
    return translate_customer_document_unified(file_type="xlsx", **kwargs)


def _require_all_candidate_customers(
    rfq_access_token: str,
    candidate_ids: list,
    *,
    require_write: bool,
) -> dict[int, str]:
    customers: dict[int, str] = {}
    for raw_id in candidate_ids:
        candidate_id = int(raw_id)
        customer_id = _candidate_customer_id(candidate_id)
        _authorize_existing_customer_unified(
            rfq_access_token,
            customer_id,
            require_write=require_write,
        )
        customers[candidate_id] = str(customer_id)
    return customers


def list_term_candidates_unified(
    *,
    rfq_access_token: str,
    customer_id: str | None = None,
    status: str | None = None,
    source_file_name: str | None = None,
    limit: int = 500,
) -> list[dict]:
    principal, allowed_ids = _confirmed_pdf_customer_ids_unified(rfq_access_token)
    if customer_id:
        _authorize_pdf_customer_unified(rfq_access_token, customer_id)
        return list_term_candidates(
            customer_id=customer_id,
            status=status,
            source_file_name=source_file_name,
            limit=limit,
        )
    if principal["role"] == "company_admin":
        return list_term_candidates(status=status, source_file_name=source_file_name, limit=limit)
    rows: list[dict] = []
    for pdf_customer_id in allowed_ids:
        rows.extend(list_term_candidates(
            customer_id=pdf_customer_id,
            status=status,
            source_file_name=source_file_name,
            limit=limit,
        ))
    return rows[: int(limit)]


def get_term_candidate_unified(candidate_id: int, *, rfq_access_token: str) -> dict | None:
    customer_id = _candidate_customer_id(int(candidate_id))
    _authorize_existing_customer_unified(rfq_access_token, customer_id)
    return get_term_candidate(int(candidate_id))


def list_term_candidates_for_job_unified(translation_job_id: str, *, rfq_access_token: str) -> dict:
    customer_id = _job_customer_id(translation_job_id)
    _authorize_existing_customer_unified(rfq_access_token, customer_id)
    return list_term_candidates_for_job(translation_job_id)


def list_candidate_occurrences_unified(candidate_id: int, *, rfq_access_token: str) -> list[dict]:
    customer_id = _candidate_customer_id(int(candidate_id))
    _authorize_existing_customer_unified(rfq_access_token, customer_id)
    return list_candidate_occurrences(int(candidate_id))


def update_term_candidate_unified(
    candidate_id: int,
    *,
    rfq_access_token: str,
    final_translation: str | None = None,
) -> dict:
    customer_id = _candidate_customer_id(int(candidate_id))
    principal, _ = _authorize_existing_customer_unified(
        rfq_access_token,
        customer_id,
        require_write=True,
    )
    return update_term_candidate(
        int(candidate_id),
        final_translation=final_translation,
        username=principal["username"],
    )


def ignore_term_candidates_unified(
    candidate_ids: list,
    *,
    rfq_access_token: str,
) -> dict:
    principal = _confirmed_pdf_customer_ids_unified(rfq_access_token)[0]
    _require_all_candidate_customers(rfq_access_token, candidate_ids, require_write=True)
    return ignore_term_candidates(candidate_ids, username=principal["username"])


def submit_term_candidates_for_approval_unified(
    rows: list[dict],
    *,
    rfq_access_token: str,
) -> dict:
    candidate_ids = [int(row.get("candidate_id")) for row in rows]
    principal = _confirmed_pdf_customer_ids_unified(rfq_access_token)[0]
    _require_all_candidate_customers(rfq_access_token, candidate_ids, require_write=True)
    sanitized_rows = [
        {
            "candidate_id": int(row.get("candidate_id")),
            "final_translation": row.get("final_translation"),
            "comment": row.get("comment", row.get("note", "")),
        }
        for row in rows
    ]
    return submit_term_candidates_for_approval(sanitized_rows, username=principal["username"])


def list_customer_glossary_unified(
    customer_id: str,
    *,
    rfq_access_token: str,
    status: str | None = "active",
    limit: int = 1000,
    offset: int = 0,
) -> list[dict]:
    _authorize_pdf_customer_unified(rfq_access_token, customer_id)
    return list_customer_glossary(customer_id, status=status, limit=limit, offset=offset)


def search_customer_glossary_unified(
    customer_id: str,
    keyword: str,
    *,
    rfq_access_token: str,
    status: str | None = "active",
    limit: int = 100,
) -> list[dict]:
    _authorize_pdf_customer_unified(rfq_access_token, customer_id)
    return search_customer_glossary(customer_id, keyword, status=status, limit=limit)


def get_glossary_term_unified(glossary_id: int, *, rfq_access_token: str) -> dict | None:
    customer_id = _glossary_customer_id(int(glossary_id))
    _authorize_existing_customer_unified(rfq_access_token, customer_id)
    return get_glossary_term(int(glossary_id))


def export_customer_glossary_unified(customer_id: str, *, rfq_access_token: str) -> bytes:
    _authorize_pdf_customer_unified(rfq_access_token, customer_id)
    return export_customer_glossary(customer_id)


def request_add_glossary_term_unified(
    *,
    rfq_access_token: str,
    customer_id: str,
    english_term: str,
    chinese_translation: str,
    note: str | None = None,
) -> dict:
    principal, _ = _authorize_pdf_customer_unified(rfq_access_token, customer_id, require_write=True)
    return request_add_glossary_term(
        customer_id=customer_id,
        english_term=english_term,
        chinese_translation=chinese_translation,
        note=note,
        submitted_by=principal["username"],
    )


def request_update_glossary_term_unified(
    *,
    rfq_access_token: str,
    glossary_id: int,
    english_term: str | None = None,
    chinese_translation: str | None = None,
    note: str | None = None,
) -> dict:
    customer_id = _glossary_customer_id(int(glossary_id))
    principal, _ = _authorize_existing_customer_unified(
        rfq_access_token,
        customer_id,
        require_write=True,
    )
    return request_update_glossary_term(
        glossary_id=int(glossary_id),
        english_term=english_term,
        chinese_translation=chinese_translation,
        note=note,
        submitted_by=principal["username"],
    )


def request_delete_glossary_term_unified(
    *,
    rfq_access_token: str,
    glossary_id: int,
    comment: str | None = None,
) -> dict:
    customer_id = _glossary_customer_id(int(glossary_id))
    principal, _ = _authorize_existing_customer_unified(
        rfq_access_token,
        customer_id,
        require_write=True,
    )
    return request_delete_glossary_term(
        glossary_id=int(glossary_id),
        submitted_by=principal["username"],
        comment=comment,
    )


def preview_customer_glossary_import_unified(
    *,
    rfq_access_token: str,
    customer_id: str,
    file_bytes: bytes,
    filename: str = "",
) -> dict:
    _authorize_pdf_customer_unified(rfq_access_token, customer_id, require_write=True)
    return preview_customer_glossary_import(
        customer_id=customer_id,
        file_bytes=file_bytes,
        filename=filename,
    )


def request_import_customer_glossary_unified(
    *,
    rfq_access_token: str,
    customer_id: str,
    file_bytes: bytes,
    filename: str = "",
) -> dict:
    principal, _ = _authorize_pdf_customer_unified(rfq_access_token, customer_id, require_write=True)
    return request_import_customer_glossary(
        customer_id=customer_id,
        file_bytes=file_bytes,
        filename=filename,
        submitted_by=principal["username"],
    )


def list_translation_jobs_unified(
    *,
    rfq_access_token: str,
    customer_id: str | None = None,
    job_type: str | None = None,
    status: str | None = None,
    limit: int = 100,
) -> list[dict]:
    principal, allowed_ids = _confirmed_pdf_customer_ids_unified(rfq_access_token)
    if customer_id:
        _authorize_pdf_customer_unified(rfq_access_token, customer_id)
        return list_translation_jobs(customer_id=customer_id, job_type=job_type, status=status, limit=limit)
    if principal["role"] == "company_admin":
        return list_translation_jobs(job_type=job_type, status=status, limit=limit)
    rows: list[dict] = []
    for pdf_customer_id in allowed_ids:
        rows.extend(list_translation_jobs(customer_id=pdf_customer_id, job_type=job_type, status=status, limit=limit))
    rows.sort(key=lambda row: row.get("created_at") or "", reverse=True)
    return rows[: int(limit)]


def get_translation_job_unified(
    job_id: str,
    *,
    rfq_access_token: str,
    include_input_bytes: bool = False,
) -> dict | None:
    customer_id = _job_customer_id(job_id)
    _authorize_existing_customer_unified(rfq_access_token, customer_id)
    return get_translation_job(job_id, include_input_bytes=include_input_bytes)


def delete_translation_job_unified(job_id: str, *, rfq_access_token: str) -> bool:
    customer_id = _job_customer_id(job_id)
    _authorize_existing_customer_unified(rfq_access_token, customer_id, require_write=True)
    return delete_translation_job(job_id)


def list_glossary_change_requests_unified(
    *,
    rfq_access_token: str,
    customer_id: str | None = None,
    status: str | None = None,
    action_type: str | None = None,
    submitted_by: str | None = None,
    limit: int | None = None,
) -> list[dict]:
    principal, allowed_ids = _confirmed_pdf_customer_ids_unified(rfq_access_token)
    if customer_id:
        _authorize_pdf_customer_unified(rfq_access_token, customer_id)
        return list_glossary_change_requests(
            customer_id=customer_id,
            status=status,
            action_type=action_type,
            submitted_by=submitted_by,
            limit=limit,
        )
    if principal["role"] == "company_admin":
        return list_glossary_change_requests(
            status=status,
            action_type=action_type,
            submitted_by=submitted_by,
            limit=limit,
        )
    rows: list[dict] = []
    for pdf_customer_id in allowed_ids:
        rows.extend(list_glossary_change_requests(
            customer_id=pdf_customer_id,
            status=status,
            action_type=action_type,
            submitted_by=submitted_by,
            limit=limit,
        ))
    rows.sort(key=lambda row: row.get("submitted_at") or "", reverse=True)
    return rows[: int(limit)] if limit is not None else rows


def get_glossary_change_request_unified(request_id: int, *, rfq_access_token: str) -> dict | None:
    customer_id = _change_request_customer_id(int(request_id))
    _authorize_existing_customer_unified(rfq_access_token, customer_id)
    return get_glossary_change_request(int(request_id))


def approve_glossary_change_request_unified(
    request_id: int,
    *,
    rfq_access_token: str,
    review_comment: str | None = None,
) -> dict:
    try:
        principal = rfq_identity.get_unified_current_user(rfq_access_token)
        if principal["role"] != "company_admin":
            raise rfq_identity.UnifiedAuthError("permission denied: reviewer is not company_admin")
        request_row = get_glossary_change_request(int(request_id))
        if request_row is None:
            return {"status": "failed", "request_id": int(request_id), "request_status": None, "error": "申请不存在"}
        if request_row.get("status") != "pending":
            return {
                "status": "failed",
                "request_id": int(request_id),
                "request_status": request_row.get("status"),
                "customer_id": request_row.get("customer_id"),
                "action_type": request_row.get("action_type"),
                "reviewed_by": request_row.get("reviewed_by"),
                "error": "request is no longer pending",
            }
        _authorize_pdf_customer_unified(rfq_access_token, request_row["customer_id"])
        reviewer = {
            "username": principal["username"],
            "role": "company_admin",
            "group_name": principal.get("group_name") or "",
            "assigned_customer_ids": [],
        }
        _core_approve_glossary_change_request(int(request_id), reviewer, review_comment or "")
        after = get_glossary_change_request(int(request_id))
        return {
            "status": "completed",
            "request_id": int(request_id),
            "request_status": (after or {}).get("status"),
            "customer_id": (after or request_row).get("customer_id"),
            "action_type": (after or request_row).get("action_type"),
            "reviewed_by": (after or {}).get("reviewed_by"),
            "reviewed_at": (after or {}).get("reviewed_at"),
            "affected_glossary_id": _affected_glossary_id_for_request(after),
            "error": None,
        }
    except Exception as exc:
        current = get_glossary_change_request(int(request_id))
        return {
            "status": "failed",
            "request_id": int(request_id),
            "request_status": (current or {}).get("status"),
            "customer_id": (current or {}).get("customer_id"),
            "action_type": (current or {}).get("action_type"),
            "reviewed_by": (current or {}).get("reviewed_by"),
            "error": str(exc),
        }


def reject_glossary_change_request_unified(
    request_id: int,
    *,
    rfq_access_token: str,
    review_comment: str | None = None,
) -> dict:
    try:
        principal = rfq_identity.get_unified_current_user(rfq_access_token)
        if principal["role"] != "company_admin":
            raise rfq_identity.UnifiedAuthError("permission denied: reviewer is not company_admin")
        request_row = get_glossary_change_request(int(request_id))
        if request_row is None:
            return {"status": "failed", "request_id": int(request_id), "request_status": None, "error": "申请不存在"}
        if request_row.get("status") != "pending":
            return {
                "status": "failed",
                "request_id": int(request_id),
                "request_status": request_row.get("status"),
                "customer_id": request_row.get("customer_id"),
                "action_type": request_row.get("action_type"),
                "reviewed_by": request_row.get("reviewed_by"),
                "error": "request is no longer pending",
            }
        _authorize_pdf_customer_unified(rfq_access_token, request_row["customer_id"])
        reviewer = {
            "username": principal["username"],
            "role": "company_admin",
            "group_name": principal.get("group_name") or "",
            "assigned_customer_ids": [],
        }
        _core_reject_glossary_change_request(int(request_id), reviewer, review_comment or "")
        after = get_glossary_change_request(int(request_id))
        return {
            "status": "completed",
            "request_id": int(request_id),
            "request_status": (after or {}).get("status"),
            "customer_id": (after or request_row).get("customer_id"),
            "action_type": (after or request_row).get("action_type"),
            "reviewed_by": (after or {}).get("reviewed_by"),
            "reviewed_at": (after or {}).get("reviewed_at"),
            "affected_glossary_id": None,
            "error": None,
        }
    except Exception as exc:
        current = get_glossary_change_request(int(request_id))
        return {
            "status": "failed",
            "request_id": int(request_id),
            "request_status": (current or {}).get("status"),
            "customer_id": (current or {}).get("customer_id"),
            "action_type": (current or {}).get("action_type"),
            "reviewed_by": (current or {}).get("reviewed_by"),
            "error": str(exc),
        }


__all__ = [
    "translate_customer_document",
    "translate_pdf_document",
    "translate_excel_document",
    "list_customers",
    "get_customer",
    "detect_pdf_translation_scope",
    "detect_excel_translation_scope",
    "list_translation_jobs",
    "get_translation_job",
    "delete_translation_job",
    "list_term_candidates",
    "get_term_candidate",
    "list_term_candidates_for_job",
    "list_candidate_occurrences",
    "update_term_candidate",
    "ignore_term_candidates",
    "submit_term_candidates_for_approval",
    "list_customer_glossary",
    "search_customer_glossary",
    "get_glossary_term",
    "export_customer_glossary",
    "request_add_glossary_term",
    "request_update_glossary_term",
    "request_delete_glossary_term",
    "preview_customer_glossary_import",
    "request_import_customer_glossary",
    "list_glossary_change_requests",
    "get_glossary_change_request",
    "approve_glossary_change_request",
    "reject_glossary_change_request",
    "translate_customer_document_unified",
    "translate_pdf_document_unified",
    "translate_excel_document_unified",
    "list_term_candidates_unified",
    "get_term_candidate_unified",
    "list_term_candidates_for_job_unified",
    "list_candidate_occurrences_unified",
    "update_term_candidate_unified",
    "ignore_term_candidates_unified",
    "submit_term_candidates_for_approval_unified",
    "list_customer_glossary_unified",
    "search_customer_glossary_unified",
    "get_glossary_term_unified",
    "export_customer_glossary_unified",
    "request_add_glossary_term_unified",
    "request_update_glossary_term_unified",
    "request_delete_glossary_term_unified",
    "preview_customer_glossary_import_unified",
    "request_import_customer_glossary_unified",
    "list_translation_jobs_unified",
    "get_translation_job_unified",
    "delete_translation_job_unified",
    "list_glossary_change_requests_unified",
    "get_glossary_change_request_unified",
    "approve_glossary_change_request_unified",
    "reject_glossary_change_request_unified",
]
