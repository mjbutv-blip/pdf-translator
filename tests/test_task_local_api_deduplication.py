from __future__ import annotations

import io
import zipfile

import fitz
import openpyxl

import translation_core


def _glossary_bytes(rows: list[tuple[str, str]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["英文术语", "中文翻译"])
    for english, chinese in rows:
        ws.append([english, chinese])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_excel_reuses_duplicate_text_and_exact_glossary_match(monkeypatch):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A20"] = "repeat sewing instruction"
    ws["A21"] = "repeat sewing instruction"
    ws["A22"] = "approved whole term"
    source = io.BytesIO()
    wb.save(source)

    calls: list[str] = []

    def fake_translate(_client, text, _glossary):
        calls.append(text)
        return "重复缝制说明"

    monkeypatch.setattr(translation_core, "OpenAI", lambda **_kwargs: object())
    monkeypatch.setattr(translation_core, "translate_cell_text", fake_translate)
    result, n_cells, _n_images, _report, _review = translation_core.run_excel_translation(
        source.getvalue(),
        _glossary_bytes([("approved whole term", "已审批完整术语")]),
        "test-key",
        lambda _text: None,
        lambda _progress: None,
    )

    translated_wb = openpyxl.load_workbook(io.BytesIO(result))
    translated_ws = translated_wb.active
    assert n_cells == 3
    assert calls == ["repeat sewing instruction"]
    assert translated_ws["A20"].value == "重复缝制说明"
    assert translated_ws["A21"].value == "重复缝制说明"
    assert translated_ws["A22"].value == "已审批完整术语"


def test_pdf_reuses_duplicate_text_across_pages(monkeypatch):
    doc = fitz.open()
    for _ in range(2):
        page = doc.new_page()
        page.insert_text((72, 300), "repeat sewing instruction")
    source = doc.tobytes()
    doc.close()

    calls: list[list[str]] = []

    def fake_batch(_client, texts, _glossary):
        calls.append(list(texts))
        return {text: "重复缝制说明" for text in texts}, set()

    monkeypatch.setattr(translation_core, "OpenAI", lambda **_kwargs: object())
    monkeypatch.setattr(translation_core, "translate_batch_resilient", fake_batch)
    translation_core.run_pdf_translation(
        source,
        _glossary_bytes([]),
        str(translation_core.DEFAULT_FONT),
        "test-key",
        lambda *_args: None,
        lambda _text: None,
        lambda _progress: None,
    )

    assert calls == [["repeat sewing instruction"]]


def test_excel_image_translation_reuses_identical_image_bytes(monkeypatch):
    source = io.BytesIO()
    with zipfile.ZipFile(source, "w") as archive:
        archive.writestr("xl/media/image1.png", b"same-image")
        archive.writestr("xl/media/image2.png", b"same-image")
        archive.writestr("xl/media/image3.png", b"different-image")

    calls: list[bytes] = []

    def fake_translate(_client, image_bytes, _ext, _glossary):
        calls.append(image_bytes)
        return b"translated:" + image_bytes

    monkeypatch.setattr(translation_core, "_translate_image_bytes", fake_translate)
    result = translation_core.translate_images_in_excel(
        source.getvalue(), object(), {}, lambda *_args: None
    )

    assert calls == [b"same-image", b"different-image"]
    with zipfile.ZipFile(io.BytesIO(result)) as archive:
        assert archive.read("xl/media/image1.png") == b"translated:same-image"
        assert archive.read("xl/media/image2.png") == b"translated:same-image"
        assert archive.read("xl/media/image3.png") == b"translated:different-image"
