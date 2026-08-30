from __future__ import annotations

import io
import json
import sys
import tempfile
import threading
import time
from datetime import datetime, timedelta
from pathlib import Path

import fitz
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import db


def _prepare_temp_db(tmpdir: str):
    db.DB_PATH = Path(tmpdir) / "excel_worker.sqlite"
    import translation_core

    translation_core.init_db()
    with db.get_db_connection() as conn:
        conn.execute(
            """
            INSERT INTO customers (
                customer_id, customer_name, customer_code, group_name,
                assigned_staff_username, created_at
            )
            VALUES ('TEST-CUST', 'Excel Worker Customer', 'TEST-CUST', 'Test', 'worker_test', ?)
            """,
            (translation_core._now_iso(),),
        )
    return translation_core


def _xlsx_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Main"
    ws["A1"] = "Hello"
    ws["B1"] = "=SUM(1,2)"
    ws.merge_cells("A2:B2")
    ws["A2"] = "Merged"
    ws2 = wb.create_sheet("Second")
    ws2["A1"] = "World"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdf_bytes() -> bytes:
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((72, 72), "Hello")
    data = doc.tobytes()
    doc.close()
    return data


def _excel_config(**overrides) -> dict:
    config = {
        "scope_mode": "all",
        "selected_sheets": None,
        "scope_detection": [],
        "translate_images": False,
    }
    config.update(overrides)
    return config


def _make_excel_job(
    filename: str = "worker.xlsx",
    *,
    execution_mode: str = "external",
    initial_status: str = "queued",
    config: dict | None = None,
) -> str:
    import translation_jobs

    if execution_mode == "external" and initial_status == "queued":
        return translation_jobs.create_external_excel_translation_job(
            username="excel_worker_test",
            customer_id="TEST-CUST",
            source_file_name=filename,
            input_bytes=_xlsx_bytes(),
            config=config or _excel_config(),
        )
    return translation_jobs.create_translation_job(
        job_type="Excel",
        username="excel_worker_test",
        customer_id="TEST-CUST",
        source_file_name=filename,
        input_bytes=_xlsx_bytes(),
        aux_bytes=None,
        config=config or _excel_config(),
        execution_mode=execution_mode,
        initial_status=initial_status,
        initial_progress=0.01 if initial_status == "running" else 0.0,
    )


def _make_pdf_job(filename: str, created_at: str) -> str:
    import translation_jobs

    job_id = translation_jobs.create_translation_job(
        job_type="PDF",
        username="excel_worker_test",
        customer_id="TEST-CUST",
        source_file_name=filename,
        input_bytes=_pdf_bytes(),
        aux_bytes=None,
        config={"scope_mode": "all"},
        execution_mode="external",
    )
    translation_jobs.update_translation_job(job_id, created_at=created_at, updated_at=created_at)
    return job_id


def _row(job_id: str) -> dict:
    import translation_jobs

    row = translation_jobs.get_translation_job_by_id(job_id)
    assert row is not None
    return row


def _fake_excel_translator(**kwargs):
    kwargs["on_cell"]("[Main] Hello")
    kwargs["on_progress"](1.0)
    return (
        kwargs["xlsx_bytes"],
        2,
        0,
        [{
            "sheet_name": "Main",
            "cell_coordinate": "A1",
            "original_text": "Hello",
            "translated_text": "你好",
            "status": "ok",
            "skip_reason": "",
            "is_merged_cell": False,
            "layout_warning": False,
            "scope_mode": kwargs.get("scope_mode", "all"),
            "selected_sheets": "",
            "skipped_sheets": "",
            "detection_score": "",
            "detection_reason": "",
        }],
        {"n_review_items": 0, "unrecorded_terms": [], "unrecorded_term_count": 0},
    )


def test_excel_job_creation() -> None:
    job_id = _make_excel_job(
        config=_excel_config(
            scope_mode="manual",
            selected_sheets=["Main"],
            scope_detection=[{"sheet_name": "Main", "is_workmanship": True}],
            translate_images=True,
        )
    )
    row = _row(job_id)
    config = json.loads(row["config"])
    assert row["job_type"] == "Excel"
    assert row["execution_mode"] == "external"
    assert row["status"] == "queued"
    assert row["input_bytes"]
    assert config["selected_sheets"] == ["Main"]
    assert config["translate_images"] is True


def test_self_contained_lifecycle_and_workbook_regression() -> None:
    import translation_jobs

    job_id = _make_excel_job("self-contained.xlsx")
    job = translation_jobs.claim_next_external_translation_job("excel-worker")
    assert job and job["job_id"] == job_id
    assert translation_jobs.run_claimed_excel_job(
        job,
        "test-key",
        "excel-worker",
        translator=_fake_excel_translator,
    ) == "complete"
    row = _row(job_id)
    assert row["status"] == "complete"
    assert row["result_file"]
    assert row["result_report"]
    meta = json.loads(row["result_meta"])
    assert meta["translated_cell_count"] == 2
    assert meta["candidate_ids_reliable"] is True
    wb = openpyxl.load_workbook(io.BytesIO(row["result_file"]), data_only=False)
    assert wb["Main"]["B1"].value == "=SUM(1,2)"
    assert "A2:B2" in [str(rng) for rng in wb["Main"].merged_cells.ranges]
    assert "Second" in wb.sheetnames


def test_excel_atomic_claim() -> None:
    import translation_jobs

    job_id = _make_excel_job("atomic.xlsx")
    barrier = threading.Barrier(2)
    results = []

    def claim(worker_id: str) -> None:
        barrier.wait()
        results.append(translation_jobs.claim_next_external_translation_job(worker_id))

    threads = [
        threading.Thread(target=claim, args=("worker-A",)),
        threading.Thread(target=claim, args=("worker-B",)),
    ]
    for thread in threads:
        thread.start()
    for thread in threads:
        thread.join()
    claimed = [job for job in results if job]
    assert len(claimed) == 1
    assert claimed[0]["job_id"] == job_id


def test_mixed_queue_fairness_by_created_at() -> None:
    import translation_core
    import translation_jobs

    base = datetime.now()
    pdf1 = _make_pdf_job("t1.pdf", (base + timedelta(seconds=1)).isoformat())
    excel = _make_excel_job("t2.xlsx")
    translation_jobs.update_translation_job(excel, created_at=(base + timedelta(seconds=2)).isoformat())
    pdf2 = _make_pdf_job("t3.pdf", (base + timedelta(seconds=3)).isoformat())

    assert translation_jobs.claim_next_external_translation_job("worker-1")["job_id"] == pdf1
    assert translation_jobs.claim_next_external_translation_job("worker-2")["job_id"] == excel
    assert translation_jobs.claim_next_external_translation_job("worker-3")["job_id"] == pdf2
    assert translation_core._now_iso()


def test_sync_isolation_and_stale_recovery() -> None:
    import translation_jobs

    stale_time = (datetime.now() - timedelta(seconds=translation_jobs.PDF_JOB_STALE_SECONDS + 10)).isoformat()
    sync_id = _make_excel_job("sync.xlsx", execution_mode="sync", initial_status="running")
    translation_jobs.update_translation_job(sync_id, worker_id=None, heartbeat_at=stale_time, updated_at=stale_time)
    assert translation_jobs.claim_next_external_translation_job("external-worker") is None
    result = translation_jobs.recover_stale_external_jobs("recovery-worker")
    assert sync_id not in result["recovered"]
    assert _row(sync_id)["status"] == "running"

    external_id = _make_excel_job("stale-excel.xlsx", execution_mode="external", initial_status="running")
    translation_jobs.update_translation_job(external_id, worker_id="lost-worker", heartbeat_at=stale_time, updated_at=stale_time, attempt_count=1)
    result = translation_jobs.recover_stale_external_jobs("recovery-worker")
    assert external_id in result["recovered"]
    assert _row(external_id)["status"] == "queued"


def test_excel_heartbeat_blocking_and_cancel() -> None:
    import translation_jobs

    original_interval = translation_jobs.PDF_JOB_HEARTBEAT_SECONDS
    translation_jobs.PDF_JOB_HEARTBEAT_SECONDS = 1
    try:
        heartbeat_id = _make_excel_job("heartbeat.xlsx")
        job = translation_jobs.claim_next_external_translation_job("heartbeat-worker")
        assert job and job["job_id"] == heartbeat_id
        old_time = "2000-01-01T00:00:00"
        translation_jobs.update_translation_job_owned(heartbeat_id, "heartbeat-worker", heartbeat_at=old_time)

        def blocking_translator(**kwargs):
            time.sleep(2.2)
            return _fake_excel_translator(**kwargs)

        thread = threading.Thread(
            target=translation_jobs.run_claimed_excel_job,
            args=(job, "test-key", "heartbeat-worker"),
            kwargs={"translator": blocking_translator},
        )
        thread.start()
        time.sleep(1.4)
        assert _row(heartbeat_id)["heartbeat_at"] != old_time
        thread.join()
        assert _row(heartbeat_id)["status"] == "complete"
    finally:
        translation_jobs.PDF_JOB_HEARTBEAT_SECONDS = original_interval

    queued_cancel_id = _make_excel_job("cancel-queued.xlsx")
    translation_jobs.cancel_excel_translation_jobs("excel_worker_test")
    assert _row(queued_cancel_id)["status"] == "failed"
    assert _row(queued_cancel_id)["error"] == translation_jobs.PDF_JOB_CANCELLED_ERROR
    assert translation_jobs.claim_next_external_translation_job("after-cancel") is None

    running_cancel_id = _make_excel_job("cancel-running.xlsx")
    job = translation_jobs.claim_next_external_translation_job("cancel-worker")
    assert job and job["job_id"] == running_cancel_id
    translation_jobs.cancel_excel_translation_jobs("excel_worker_test")
    assert translation_jobs.run_claimed_excel_job(
        job,
        "test-key",
        "cancel-worker",
        translator=lambda **kwargs: (time.sleep(0.1) or _fake_excel_translator(**kwargs)),
    ) == "cancelled"
    assert _row(running_cancel_id)["result_file"] is None


def test_excel_ownership_protection() -> None:
    import translation_jobs

    job_id = _make_excel_job("ownership.xlsx")
    job = translation_jobs.claim_next_external_translation_job("worker-A")
    assert job and job["job_id"] == job_id
    translation_jobs.update_translation_job(job_id, status="queued", worker_id=None, heartbeat_at=None)
    reclaimed = translation_jobs.claim_next_external_translation_job("worker-B")
    assert reclaimed and reclaimed["job_id"] == job_id
    assert not translation_jobs.update_translation_job_owned(job_id, "worker-A", status="complete", result_file=b"stale")
    row = _row(job_id)
    assert row["status"] == "running"
    assert row["worker_id"] == "worker-B"
    assert row["result_file"] is None


def test_image_translation_branch_with_fake_image_translator() -> None:
    import translation_jobs

    job_id = _make_excel_job("images.xlsx", config=_excel_config(translate_images=True))
    job = translation_jobs.claim_next_external_translation_job("image-worker")
    assert job and job["job_id"] == job_id
    calls = {"image": 0}

    def fake_image_translator(xlsx_bytes, _client, _glossary, on_image, selected_sheets=None):
        calls["image"] += 1
        on_image(1, 1, "image1.png")
        return xlsx_bytes, [{
            "drawing": "drawing1.xml",
            "image": "image1.png",
            "status": "ok",
            "original_text": "HELLO",
            "translated_text": "你好",
            "skip_reason": "",
        }]

    assert translation_jobs.run_claimed_excel_job(
        job,
        "test-key",
        "image-worker",
        translator=_fake_excel_translator,
        image_translator=fake_image_translator,
    ) == "complete"
    assert calls["image"] == 1
    meta = json.loads(_row(job_id)["result_meta"])
    assert meta["translate_images"] is True
    assert meta["translated_image_count"] == 1


def main() -> None:
    with tempfile.TemporaryDirectory() as tmp:
        _prepare_temp_db(tmp)
        test_excel_job_creation()
    with tempfile.TemporaryDirectory() as tmp:
        _prepare_temp_db(tmp)
        test_self_contained_lifecycle_and_workbook_regression()
    with tempfile.TemporaryDirectory() as tmp:
        _prepare_temp_db(tmp)
        test_excel_atomic_claim()
    with tempfile.TemporaryDirectory() as tmp:
        _prepare_temp_db(tmp)
        test_mixed_queue_fairness_by_created_at()
    with tempfile.TemporaryDirectory() as tmp:
        _prepare_temp_db(tmp)
        test_sync_isolation_and_stale_recovery()
    with tempfile.TemporaryDirectory() as tmp:
        _prepare_temp_db(tmp)
        test_excel_heartbeat_blocking_and_cancel()
    with tempfile.TemporaryDirectory() as tmp:
        _prepare_temp_db(tmp)
        test_excel_ownership_protection()
    with tempfile.TemporaryDirectory() as tmp:
        _prepare_temp_db(tmp)
        test_image_translation_branch_with_fake_image_translator()
    print("excel worker job tests passed")


if __name__ == "__main__":
    main()
